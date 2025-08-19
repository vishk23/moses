"""
Balance Tracker - revised
Developed by CD
[v?]
"""
from pathlib import Path

import pandas as pd # type: ignore
import openpyxl # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

# Template and new data loaded in
df1_path = Path('./assets/Portfolio_Balance_Tracker_2025YTD.xlsx') 
df1 = pd.read_excel(df1_path,engine='openpyxl')
df2 = pd.read_excel(Path(r'H:\FinishedReports\BalanceTracker\Source\Summary for Portfolio Balance Tracker_2025-01-31.xlsx')) 


def update_excel_template(template_path, new_data_df, output_path):
    wb = openpyxl.load_workbook(template_path)
    sheet = wb["Calendar 2025"]
    
    last_filled_month_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=5, column=col).value is None:
            last_filled_month_col = col - 1
            break
    
    for i, value in enumerate(new_data_df.iloc[:, 0]):
        row = 5 + (i * 2)
        sheet.cell(row=row, column=last_filled_month_col + 1).value = value
        
    wb.save(output_path)
    
output_path = df1_path

update_excel_template(df1_path,df2,output_path)

