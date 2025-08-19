import openpyxl # type: ignore
from openpyxl.utils import get_column_letter # type: ignore


def update_excel_template(template_path, df, additional_fields, output_path):
    wb = openpyxl.load_workbook(template_path)
    sheet = wb["Calendar 2025"]

    category_to_row = {}
    for row in sheet.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value is not None:
            category_to_row[cell.value] = cell.row

    for index, record in df.iterrows():
        category = record['Category']
        delta = record['Delta']

        if category not in category_to_row:
            print(f"Warning: Category '{category}' not in column A. Skipping.")
            continue

        row_number = category_to_row[category]
        col_number = 3

        while col_number <= 14: # Stop after December written
            current_cell = sheet.cell(row=row_number, column=col_number)
            if current_cell.value is None:
                current_cell.value = delta
                break
            col_number += 1
        else:
            print(f"Error, Calendar completed. Please refresh for a new year in a new template tab.")

    for index, row in additional_fields.iterrows():
        category = row['Category']
        if category in category_to_row:
            excel_row = category_to_row[category]
            # Update columns Q, R, S (17, 18, 19)
            sheet.cell(row=excel_row, column=17).value = row['New Loan Yield']
            sheet.cell(row=excel_row, column=18).value = row['Total Loan Yield']
            sheet.cell(row=excel_row, column=19).value = row['Unadvanced Funds']
    
    wb.save(output_path)
    wb.close()