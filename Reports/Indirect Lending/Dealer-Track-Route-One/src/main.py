"""
Dealer Track - Route One Reconciliation Report - Main Entry Point

BUSINESS LOGIC EXTRACTED:

[DEBUG] Headers after reading dtvault:
  Columns: ['Container ID', 'Creation Date', 'Status', 'Vault ID', 'Billing Type', 'Asset Type', 'Transaction Creator', 'Customer Name', 'Dealership Name', 'Loan Amount (cents)']...
  Shape: (92, 16)

[DEBUG] Headers after reading funding:
  Columns: ['Application Number', 'Application Date', 'Application Status', 'Applicant Name', 'Co-Applicant Name', 'Portal', 'Dealer Name', 'Effective Fund Date', 'Funded by', 'Dealer Flat']...
  Shape: (319, 19)

[DEBUG] Headers after reading routeonevault:
  Columns: ['Assignee of Record', 'Dealership Name', 'VIN', 'Customer Name', 'Contract Date', 'FS Acceptance Date', 'Sales Class', 'Trans. Type', 'Product Type', 'R1 Contract Number']...
  Shape: (193, 18)

[DEBUG] Headers after splitting:
  Funding columns: ['Application Number', 'Application Date', 'Application Status', 'Applicant Name', 'Co-Applicant Name', 'Portal', 'Dealer Name', 'Effective Fund Date', 'Funded by', 'Dealer Flat']...
  RouteOne columns: ['Assignee of Record', 'Dealership Name', 'VIN', 'Customer Name', 'Contract Date', 'FS Acceptance Date', 'Sales Class', 'Trans. Type', 'Product Type', 'R1 Contract Number']...
  DT Vault columns: ['Container ID', 'Creation Date', 'Status', 'Vault ID', 'Billing Type', 'Asset Type', 'Transaction Creator', 'Customer Name', 'Dealership Name', 'Loan Amount (cents)']...

[DEBUG] Headers after cleanup (before adding Reconciled column):
  Remaining funding rows: 56
  Remaining routeone rows: 1
  Remaining dtvault rows: 1
  RouteOne columns: ['Assignee of Record', 'Dealership Name', 'VIN', 'Customer Name', 'Contract Date', 'FS Acceptance Date', 'Sales Class', 'Trans. Type', 'Product Type', 'R1 Contract Number']...
  DT Vault columns: ['Container ID', 'Creation Date', 'Status', 'Vault ID', 'Billing Type', 'Asset Type', 'Transaction Creator', 'Customer Name', 'Dealership Name', 'Loan Amount (cents)']...

[DEBUG] Unmatched funding headers:
  Unmatched RouteOne Econtracts: 0 rows
  Unmatched DT Econtracts: 2 rows
    Columns: ['Application Number', 'Application Date', 'Application Status', 'Applicant Name', 'Co-Applicant Name', 'Portal', 'Dealer Name', 'Effective Fund Date', 'Funded by', 'Dealer Flat']...

[DEBUG] DT Vault headers for NOT RECONCILED (Section 1):
  Original headers: ['Container ID', 'Reconciled', 'Creation Date', 'Status', 'Vault ID', 'Billing Type', 'Asset Type', 'Transaction Creator', 'Customer Name', 'Dealership Name']...

[DEBUG] DT Funding headers for NOT RECONCILED (Section 2):
  Original headers: ['Application Number', 'Application Date', 'Application Status', 'Applicant Name', 'Co-Applicant Name', 'Portal', 'Dealer Name', 'Effective Fund Date', 'Funded by', 'Dealer Flat']...

[DEBUG] RouteOne Vault headers for NOT RECONCILED (Section 1):
  Original headers: ['Assignee of Record', 'Reconciled', 'Dealership Name', 'VIN', 'Customer Name', 'Contract Date', 'FS Acceptance Date', 'Sales Class', 'Trans. Type', 'Product Type']...

[DEBUG] Final NOT RECONCILED structured dataframes:
  DT NOT RECONCILED shape: (8, 20)
    First 3 rows of data:
      Row 0: IN VAULT BUT NOT IN FUNDING...
      Row 1: Container ID...
      Row 2: 3655030501-310200026950074660_130308_130308_DT_BCS...
  RouteOne NOT RECONCILED shape: (6, 20)
    First 3 rows of data:
      Row 0: IN VAULT BUT NOT IN FUNDING...
      Row 1: Assignee of Record...
      Row 2: Bristol County Savings Bank...
Data Sources & Relationships:
- Funding reports containing application numbers and amounts
- Route One Vault data with loan applications and FS acceptance dates
- Dealer Track Vault data with DT-formatted application IDs
- Three input files required: dt vault, funding, routeonevault

Business Rules:
- Files must contain month/year identifiers for validation
- Route One vault records filtered by FS Acceptance Date matching target month/year
- Application number matching: RouteOne uses direct match, DT uses pattern matching
- Amount validation required for successful reconciliation
- Unmatched records categorized by contract type (E Contract vs Paper)

Data Processing Flow:
1. Validate input files (exactly 3 required with specific keywords)
2. Extract month/year from filenames and validate consistency
3. Load and preprocess data (headers, datetime conversion, string cleaning)
4. Filter Route One data by FS Acceptance Date for target month
5. Match funding records to vault data (RouteOne direct, DT pattern-based)
6. Categorize results: reconciled, errors, not reconciled, missing
7. Generate summary statistics and verification totals
8. Create formatted Excel output with multiple sheets
9. Archive processed files and clean up input directory

Key Calculations:
- RouteOne reconciliation: direct application number match + amount validation
- DT reconciliation: pattern matching (appnum_DT_BCS) + amount validation
- Summary totals: reconciled, not reconciled, missing, errors by system
- Verification: total processed should equal original funding records

Business Intelligence Value:
- Ensures data integrity between lending systems
- Identifies missing or mismatched loan records
- Tracks electronic contract processing accuracy
- Supports monthly reconciliation processes
- Provides audit trail for regulatory compliance
"""

import os
from pathlib import Path
import pandas as pd
import re
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl import styles
from shutil import copy2
from datetime import datetime
import src.config

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=FutureWarning)

def main():
    """Main report execution function"""
    env = os.getenv('REPORT_ENV', 'dev')
    
    BASE_PATH = src.config.BASE_PATH 
    print(BASE_PATH)
    for item in BASE_PATH.iterdir():
        print(item)
   
    INPUT_PATH: Path = BASE_PATH / Path('./Input')
    OUTPUT_PATH: Path = BASE_PATH / Path('./Reports')
    ARCHIVE_PATH: Path = BASE_PATH / Path('./Archive')
    # Input path
    print('---------')
    print(INPUT_PATH)
    for item in INPUT_PATH.iterdir():
        print(item)
 
    expected_keywords = ["dtvault", "funding", "routeonevault"]
    month_regex = r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|" \
                r"january|february|march|april|may|june|july|august|september|october|november|december)"
    year_regex = r"\b\d{4}\b"

    files = [f for f in os.listdir(INPUT_PATH) if f.endswith(('.xlsx', '.xls'))]
    assert len(files) == 3, f"Expected 3 files, found {len(files)}"

    matched_files = {}
    month_years = set()

    for keyword in expected_keywords:
        found = False
        for f in files:
            filepath_name = f
            f_date_regex = f.lower()
            print(f)
            f = f.lower().replace(" ", "")
            if keyword in f:
                print(keyword)
                month_match = re.search(month_regex, f_date_regex)
                year_match = re.search(year_regex, f_date_regex)
                if month_match and year_match:
                    month_year = f"{month_match.group()} {year_match.group()}"
                    month_years.add(month_year)
                    matched_files[keyword] = os.path.join(INPUT_PATH, filepath_name)
                    found = True
                    break
        if not found:
            raise FileNotFoundError(f"No matching file found for keyword '{keyword}' with valid month and year.")

    assert len(month_years) == 1, f"Files do not have matching month and year values: {month_years}"
    month_year = month_years.pop()
    output_path = OUTPUT_PATH / f"{month_year} DL Reconciliation Report.xlsx"

    # Utility to drop rows where all values except 'Reconciled' are empty/NaT/nan
    def drop_blank_rows_except_reconciled(df):
        non_rec_cols = df.drop(columns=["Reconciled"], errors="ignore")
        is_blank = non_rec_cols.applymap(lambda x: str(x).strip() in ["", "nan", "NaT"])
        return df[~is_blank.all(axis=1)]

    # Load all 3 files
    dfs = {}
    for key, path in matched_files.items():
        df = pd.read_excel(path, header=2 if key == "funding" else 0)
        
        # DEBUG: Print headers after initial read
        print(f"\n[DEBUG] Headers after reading {key}:")
        print(f"  Columns: {list(df.columns)[:10]}...")  # Show first 10 columns
        print(f"  Shape: {df.shape}")

        if key == "funding":
            # Handle None values before datetime conversion
            df.iloc[:, 1] = df.iloc[:, 1].fillna(pd.NaT)
            try:
                df.iloc[:, 1] = pd.to_datetime(df.iloc[:, 1], errors="coerce")
            except Exception as e:
                print(f"Warning: Could not parse datetime column 1 in {key}: {e}")
                df.iloc[:, 1] = pd.NaT
        elif key == "dtevault":
            # Handle None values before datetime conversion
            df.iloc[:, 1] = df.iloc[:, 1].fillna(pd.NaT)
            try:
                df.iloc[:, 1] = pd.to_datetime(df.iloc[:, 1], errors="coerce")
            except Exception as e:
                print(f"Warning: Could not parse datetime column 1 in {key}: {e}")
                df.iloc[:, 1] = pd.NaT
        elif key == "routeonevault" :
            # Handle None values before datetime conversion
            df.iloc[:, 4] = df.iloc[:, 4].fillna(pd.NaT)
            df.iloc[:, 5] = df.iloc[:, 5].fillna(pd.NaT)
            try:
                df.iloc[:, 4] = pd.to_datetime(df.iloc[:, 4], errors="coerce")
                df.iloc[:, 5] = pd.to_datetime(df.iloc[:, 5], errors="coerce")
            except Exception as e:
                print(f"Warning: Could not parse datetime columns 4,5 in {key}: {e}")
                df.iloc[:, 4] = pd.NaT
                df.iloc[:, 5] = pd.NaT
        for i in range(df.shape[1]):
            if key == "funding" and i == 1:
                continue
            if key == "dtevault" and i == 1:
                continue
            if key == "routeonevault" and i in [4, 5]:
                continue
            try:
                df.iloc[:, i] = df.iloc[:, i].astype(str).replace("nan", "")
            except:
                pass
        dfs[key] = df

    # Split into components
    funding_df = dfs["funding"]
    routeone_df = dfs["routeonevault"]
    dtvault_df = dfs["dtvault"]
    
    # DEBUG: Print headers after splitting
    print(f"\n[DEBUG] Headers after splitting:")
    print(f"  Funding columns: {list(funding_df.columns)[:10]}...")
    print(f"  RouteOne columns: {list(routeone_df.columns)[:10]}...")
    print(f"  DT Vault columns: {list(dtvault_df.columns)[:10]}...")
    
    # Filter Route One Vault records to only include those with FS Acceptance Date in the current month
    # Parse the month_year to get the target month and year
    month_map = {
        'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
        'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
        'aug': 8, 'august': 8, 'sep': 9, 'september': 9, 'oct': 10, 'october': 10,
        'nov': 11, 'november': 11, 'dec': 12, 'december': 12
    }
    
    month_part = month_year.split()[0].lower()
    year_part = int(month_year.split()[1])
    target_month = month_map[month_part]
    target_year = year_part
    

    
    # Filter routeone_df to only include records where FS Acceptance Date is in the target month/year
    if not routeone_df.empty:
        # Create a mask for records in the target month/year
        fs_acceptance_dates = pd.to_datetime(routeone_df.iloc[:, 5], errors='coerce')
        month_year_mask = (fs_acceptance_dates.dt.month == target_month) & (fs_acceptance_dates.dt.year == target_year)
        
        # Apply the filter - use .copy() to avoid SettingWithCopyWarning
        original_count = len(routeone_df)
        routeone_df = routeone_df[month_year_mask].copy()
        filtered_count = len(routeone_df)

        
        # Reset index after filtering
        routeone_df.reset_index(drop=True, inplace=True)

    # Match logic
    routeone_resolved, dt_resolved, routeone_errors, dt_errors = [], [], [], []
    funding_indices_to_drop, routeone_indices_to_drop, dtvault_indices_to_drop = [], [], []
    
    # Debug counters
    routeone_matches_found = 0
    dt_matches_found = 0

    # Debug: Print some sample data to understand the matching issue

    
    # Debug: Check RouteOne Econtracts in funding data
    routeone_funding = funding_df[funding_df.iloc[:, 5] == 'Route One']  # Portal column

    
    # Debug: Check which RouteOne Econtracts are missing from vault
    routeone_app_nums = set(routeone_funding.iloc[:, 0].astype(str))
    vault_app_nums = set(routeone_df.iloc[:, 15].astype(str).str.replace('.0', ''))
    missing_from_vault = routeone_app_nums - vault_app_nums

    
    for i, fund_row in funding_df.iterrows():
        app_num = str(fund_row.iloc[0]).strip()
        fund_amt = fund_row.iloc[12]

        matched_ro = routeone_df[routeone_df.iloc[:, 15].astype(str).str.strip().str.replace('.0', '') == app_num]
        if not matched_ro.empty:
            routeone_matches_found += 1
            for j, ro_row in matched_ro.iterrows():
                ro_amt = ro_row.iloc[16]
                if pd.notna(ro_amt) and pd.notna(fund_amt) and float(ro_amt) == float(fund_amt):
                    combined = pd.concat([fund_row, ro_row]).to_frame().T # type: ignore
                    combined.insert(1, "Reconciled", "Yes")
                    routeone_resolved.append(combined)
                    funding_indices_to_drop.append(i)
                    routeone_indices_to_drop.append(j)
                    break
                else:
                    error_row = pd.concat([fund_row, ro_row]).to_frame().T # type: ignore
                    routeone_errors.append(error_row)
                    funding_indices_to_drop.append(i)
                    routeone_indices_to_drop.append(j)
                    break
            continue

        pattern = f"{app_num}_DT_BCS"
        matched_dt = dtvault_df[dtvault_df.iloc[:, 0].astype(str).str.endswith(pattern)]
        if not matched_dt.empty:
            dt_matches_found += 1
            for j, dt_row in matched_dt.iterrows():
                dt_amt = dt_row.iloc[9]
                if pd.notna(dt_amt) and pd.notna(fund_amt) and float(dt_amt) == float(fund_amt):
                    combined = pd.concat([fund_row, dt_row]).to_frame().T # type: ignore
                    combined.insert(1, "Reconciled", "Yes")
                    dt_resolved.append(combined)
                    funding_indices_to_drop.append(i)
                    dtvault_indices_to_drop.append(j)
                    break
                else:
                    error_row = pd.concat([fund_row, dt_row]).to_frame().T # type: ignore
                    dt_errors.append(error_row)
                    funding_indices_to_drop.append(i)
                    dtvault_indices_to_drop.append(j)
                    break
            continue

    # Build resolved/error DFs

    
    routeone_resolved_df = pd.concat(routeone_resolved, ignore_index=True) if routeone_resolved else pd.DataFrame()
    dt_resolved_df = pd.concat(dt_resolved, ignore_index=True) if dt_resolved else pd.DataFrame()
    routeone_error_df = pd.concat(routeone_errors, ignore_index=True) if routeone_errors else pd.DataFrame()
    dt_error_df = pd.concat(dt_errors, ignore_index=True) if dt_errors else pd.DataFrame()

    # Unmatched cleanup
    funding_df.drop(index=funding_indices_to_drop, inplace=True)
    routeone_df.drop(index=routeone_indices_to_drop, inplace=True)
    dtvault_df.drop(index=dtvault_indices_to_drop, inplace=True)
    routeone_df.reset_index(drop=True, inplace=True)
    dtvault_df.reset_index(drop=True, inplace=True)
    
    # DEBUG: Print headers after cleanup
    print(f"\n[DEBUG] Headers after cleanup (before adding Reconciled column):")
    print(f"  Remaining funding rows: {len(funding_df)}")
    print(f"  Remaining routeone rows: {len(routeone_df)}")
    print(f"  Remaining dtvault rows: {len(dtvault_df)}")
    if not routeone_df.empty:
        print(f"  RouteOne columns: {list(routeone_df.columns)[:10]}...")
    if not dtvault_df.empty:
        print(f"  DT Vault columns: {list(dtvault_df.columns)[:10]}...")

    # Insert reconciled = No + drop blanks
    if not routeone_df.empty:
        routeone_df.insert(1, "Reconciled", "No")
        routeone_df = drop_blank_rows_except_reconciled(routeone_df)

    if not dtvault_df.empty:
        dtvault_df.insert(1, "Reconciled", "No")
        dtvault_df = drop_blank_rows_except_reconciled(dtvault_df)

    # Separate unmatched funding records by type
    # RouteOne Econtracts that couldn't be matched to vault
    unmatched_routeone_econtracts = funding_df[(funding_df.iloc[:, 5] == 'Route One') & (funding_df.iloc[:, 15] == 'E Contract')]
    
    # DT Econtracts that couldn't be matched to vault  
    unmatched_dt_econtracts = funding_df[(funding_df.iloc[:, 5] != 'Route One') & (funding_df.iloc[:, 15] == 'E Contract')]
    
    # All Paper contracts (both RouteOne and DT)
    unmatched_paper_contracts = funding_df[funding_df.iloc[:, 15] == 'Paper']
    
    # DEBUG: Print unmatched headers
    print(f"\n[DEBUG] Unmatched funding headers:")
    print(f"  Unmatched RouteOne Econtracts: {len(unmatched_routeone_econtracts)} rows")
    if not unmatched_routeone_econtracts.empty:
        print(f"    Columns: {list(unmatched_routeone_econtracts.columns)[:10]}...")
    print(f"  Unmatched DT Econtracts: {len(unmatched_dt_econtracts)} rows")
    if not unmatched_dt_econtracts.empty:
        print(f"    Columns: {list(unmatched_dt_econtracts.columns)[:10]}...")
    

    
    # Calculate summary data and print to command line
    total_funding_records = len(funding_df) + len(funding_indices_to_drop)
    total_routeone_vault_original = len(routeone_df) + len(routeone_indices_to_drop)
    total_dt_vault_original = len(dtvault_df) + len(dtvault_indices_to_drop)
    
    # Calculate all subtotals in Python
    routeone_reconciled = len(routeone_resolved)
    routeone_not_reconciled = len(routeone_df)
    routeone_missing = len(unmatched_routeone_econtracts)
    routeone_errors_count = len(routeone_errors)
    routeone_subtotal = routeone_reconciled + routeone_not_reconciled + routeone_missing + routeone_errors_count
    
    dt_reconciled = len(dt_resolved)
    dt_not_reconciled = len(dtvault_df)
    dt_missing = len(unmatched_dt_econtracts)
    dt_errors_count = len(dt_errors)
    dt_subtotal = dt_reconciled + dt_not_reconciled + dt_missing + dt_errors_count
    
    paper_total = len(unmatched_paper_contracts)
    total_processed = routeone_subtotal + dt_subtotal + paper_total
    
    # Print summary to command line

    
    # Hardcode summary data - no DataFrame creation
    summary_categories = [
        '=== INPUT RECORDS ===',
        'Funding Records (Total)',
        'RouteOne Vault Records (After Filtering)',
        'DT Vault Records (Total)',
        '',
        '=== ROUTEONE ECONTRACTS ===',
        '  Reconciled (Matched + Amount Match)',
        '  Not Reconciled (In Vault, No Match)',
        '  Missing from Vault (In Funding Only)',
        '  Errors (Matched, Amount Mismatch)',
        '  RouteOne Econtracts Subtotal',
        '',
        '=== DT ECONTRACTS ===',
        '  Reconciled (Matched + Amount Match)',
        '  Not Reconciled (In Vault, No Match)',
        '  Missing from Vault (In Funding Only)',
        '  Errors (Matched, Amount Mismatch)',
        '  DT Econtracts Subtotal',
        '',
        '=== PAPER CONTRACTS ===',
        '  Paper Contracts (All Types)',
        '',
        '=== VERIFICATION ===',
        '  Total Processed (Should Equal Funding Total)',
        '  Funding Total (Original)'
    ]
    
    summary_counts = [
        '',
        total_funding_records,
        total_routeone_vault_original,
        total_dt_vault_original,
        '',
        '',
        routeone_reconciled,
        routeone_not_reconciled,
        routeone_missing,
        routeone_errors_count,
        routeone_subtotal,
        '',
        '',
        dt_reconciled,
        dt_not_reconciled,
        dt_missing,
        dt_errors_count,
        dt_subtotal,
        '',
        '',
        paper_total,
        '',
        '',
        total_processed,
        total_funding_records
    ]
    
    # Build final export dict with month/year suffix
    suffix = f"({month_year})"
    
    # For Route One Vault sheet, sort to put Paper contracts at the bottom
    routeone_vault_df = routeone_resolved_df.copy() if not routeone_resolved_df.empty else pd.DataFrame()
    if not routeone_vault_df.empty:
        # Check if column Q (17th column, index 16) exists (contract type column)
        if routeone_vault_df.shape[1] > 16:
            # Create a sort key: Paper contracts get value 1, others get 0
            sort_key = (routeone_vault_df.iloc[:, 16] == 'Paper').astype(int)
            # Sort by this key to put Paper contracts at the bottom
            routeone_vault_df = routeone_vault_df.iloc[sort_key.argsort()]
            routeone_vault_df.reset_index(drop=True, inplace=True)
    
    # Create structured NOT RECONCILED dataframes with sections
    # For DT Vault NOT RECONCILED - Always show both headers
    dt_not_reconciled_structured = pd.DataFrame()
    all_rows = []
    
    # Determine the maximum number of columns needed
    # Use a minimum of 20 columns to ensure proper structure even if data is empty
    # Subtract 1 from vault df columns to account for "Reconciled" column removal
    max_cols = max(
        (dtvault_df.shape[1] - 1) if not dtvault_df.empty else 20,  # -1 for Reconciled column
        unmatched_dt_econtracts.shape[1] if not unmatched_dt_econtracts.empty else 20,
        20  # Minimum columns to ensure structure
    )
    
    # Section 1: In Vault but not in Funding - ALWAYS show header
    # Add section header row
    all_rows.append(["IN VAULT BUT NOT IN FUNDING"] + [""] * (max_cols - 1))
    
    if not dtvault_df.empty:
        # Remove the "Reconciled" column for consistency with funding headers
        dtvault_for_export = dtvault_df.drop(columns=["Reconciled"], errors="ignore")
        # Add the actual column headers from dtvault_df (without Reconciled)
        vault_headers = list(dtvault_for_export.columns)
        # DEBUG: Print vault headers before padding
        print(f"\n[DEBUG] DT Vault headers for NOT RECONCILED (Section 1):")
        print(f"  Original headers: {vault_headers[:10]}...")
        # Pad with empty strings if needed
        while len(vault_headers) < max_cols:
            vault_headers.append("")
        all_rows.append(vault_headers)
        
        # Add the data rows
        for _, row in dtvault_for_export.iterrows():
            row_data = list(row)
            while len(row_data) < max_cols:
                row_data.append("")
            all_rows.append(row_data)
    else:
        # Even if empty, add a placeholder row to indicate no data
        all_rows.append(["No data in this section"] + [""] * (max_cols - 1))
    
    # Add separator row between sections
    all_rows.append([""] * max_cols)  # Empty separator row
    
    # Section 2: In Funding but not in Vault - ALWAYS show header
    # Add section header row
    all_rows.append(["IN FUNDING BUT NOT IN VAULT"] + [""] * (max_cols - 1))
    
    if not unmatched_dt_econtracts.empty:
        # Add the actual column headers from unmatched_dt_econtracts
        funding_headers = list(unmatched_dt_econtracts.columns)
        # DEBUG: Print funding headers before padding
        print(f"\n[DEBUG] DT Funding headers for NOT RECONCILED (Section 2):")
        print(f"  Original headers: {funding_headers[:10]}...")
        # Pad with empty strings if needed
        while len(funding_headers) < max_cols:
            funding_headers.append("")
        all_rows.append(funding_headers)
        
        # Add the data rows
        for _, row in unmatched_dt_econtracts.iterrows():
            row_data = list(row)
            while len(row_data) < max_cols:
                row_data.append("")
            all_rows.append(row_data)
    else:
        # Even if empty, add a placeholder row to indicate no data
        all_rows.append(["No data in this section"] + [""] * (max_cols - 1))
    
    # Create DataFrame from all rows
    # Create generic column names for the structured dataframe
    generic_columns = [f"Column_{i+1}" for i in range(max_cols)]
    dt_not_reconciled_structured = pd.DataFrame(all_rows, columns=generic_columns)
    
    # For Route One Vault NOT RECONCILED - Always show both headers
    routeone_not_reconciled_structured = pd.DataFrame()
    all_rows = []
    
    # Determine the maximum number of columns needed
    # Use a minimum of 20 columns to ensure proper structure even if data is empty
    # Subtract 1 from vault df columns to account for "Reconciled" column removal
    max_cols = max(
        (routeone_df.shape[1] - 1) if not routeone_df.empty else 20,  # -1 for Reconciled column
        unmatched_routeone_econtracts.shape[1] if not unmatched_routeone_econtracts.empty else 20,
        20  # Minimum columns to ensure structure
    )
    
    # Section 1: In Vault but not in Funding - ALWAYS show header
    # Add section header row
    all_rows.append(["IN VAULT BUT NOT IN FUNDING"] + [""] * (max_cols - 1))
    
    if not routeone_df.empty:
        # Remove the "Reconciled" column for consistency with funding headers
        routeone_for_export = routeone_df.drop(columns=["Reconciled"], errors="ignore")
        # Add the actual column headers from routeone_df (without Reconciled)
        vault_headers = list(routeone_for_export.columns)
        # DEBUG: Print vault headers before padding
        print(f"\n[DEBUG] RouteOne Vault headers for NOT RECONCILED (Section 1):")
        print(f"  Original headers: {vault_headers[:10]}...")
        # Pad with empty strings if needed
        while len(vault_headers) < max_cols:
            vault_headers.append("")
        all_rows.append(vault_headers)
        
        # Add the data rows
        for _, row in routeone_for_export.iterrows():
            row_data = list(row)
            while len(row_data) < max_cols:
                row_data.append("")
            all_rows.append(row_data)
    else:
        # Even if empty, add a placeholder row to indicate no data
        all_rows.append(["No data in this section"] + [""] * (max_cols - 1))
    
    # Add separator row between sections
    all_rows.append([""] * max_cols)  # Empty separator row
    
    # Section 2: In Funding but not in Vault - ALWAYS show header
    # Add section header row
    all_rows.append(["IN FUNDING BUT NOT IN VAULT"] + [""] * (max_cols - 1))
    
    if not unmatched_routeone_econtracts.empty:
        # Add the actual column headers from unmatched_routeone_econtracts
        funding_headers = list(unmatched_routeone_econtracts.columns)
        # DEBUG: Print funding headers before padding
        print(f"\n[DEBUG] RouteOne Funding headers for NOT RECONCILED (Section 2):")
        print(f"  Original headers: {funding_headers[:10]}...")
        # Pad with empty strings if needed
        while len(funding_headers) < max_cols:
            funding_headers.append("")
        all_rows.append(funding_headers)
        
        # Add the data rows
        for _, row in unmatched_routeone_econtracts.iterrows():
            row_data = list(row)
            while len(row_data) < max_cols:
                row_data.append("")
            all_rows.append(row_data)
    else:
        # Even if empty, add a placeholder row to indicate no data
        all_rows.append(["No data in this section"] + [""] * (max_cols - 1))
    
    # Create DataFrame from all rows
    # Create generic column names for the structured dataframe
    generic_columns = [f"Column_{i+1}" for i in range(max_cols)]
    routeone_not_reconciled_structured = pd.DataFrame(all_rows, columns=generic_columns)
    
    # DEBUG: Print final structured dataframe info
    print(f"\n[DEBUG] Final NOT RECONCILED structured dataframes:")
    print(f"  DT NOT RECONCILED shape: {dt_not_reconciled_structured.shape}")
    if not dt_not_reconciled_structured.empty:
        print(f"    First 3 rows of data:")
        for i in range(min(3, len(dt_not_reconciled_structured))):
            print(f"      Row {i}: {dt_not_reconciled_structured.iloc[i, 0][:50]}...")
    
    print(f"  RouteOne NOT RECONCILED shape: {routeone_not_reconciled_structured.shape}")
    if not routeone_not_reconciled_structured.empty:
        print(f"    First 3 rows of data:")
        for i in range(min(3, len(routeone_not_reconciled_structured))):
            print(f"      Row {i}: {routeone_not_reconciled_structured.iloc[i, 0][:50]}...")
    
    dfs_to_export = {
        f"Route One Vault {suffix}": routeone_vault_df,
        f"Route One Vault NOT RECONCILED {suffix}": routeone_not_reconciled_structured,
        f"Route One Vault Errors {suffix}": routeone_error_df,
        f"DT Vault {suffix}": dt_resolved_df,
        f"DT Vault NOT RECONCILED {suffix}": dt_not_reconciled_structured,
        f"DT Vault Errors {suffix}": dt_error_df,
        f"{month_year} DL Paper contract Report": unmatched_paper_contracts,
    }

    # Create new workbook manually - no pandas ExcelWriter
    wb = Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Create Summary sheet first - write data in most basic way
    summary_sheet_name = f"Summary {suffix}"
    ws_summary = wb.create_sheet(title=summary_sheet_name[:31])
    
    # Write summary data one cell at a time - no loops
    ws_summary.cell(row=1, column=1, value="Category")
    ws_summary.cell(row=1, column=2, value="Count")
    
    ws_summary.cell(row=2, column=1, value="INPUT RECORDS ")
    ws_summary.cell(row=3, column=1, value="Funding Records (Total)")
    ws_summary.cell(row=3, column=2, value=str(total_funding_records))
    ws_summary.cell(row=4, column=1, value="RouteOne Vault Records (After Filtering)")
    ws_summary.cell(row=4, column=2, value=str(total_routeone_vault_original))
    ws_summary.cell(row=5, column=1, value="DT Vault Records (Total)")
    ws_summary.cell(row=5, column=2, value=str(total_dt_vault_original))
    
    ws_summary.cell(row=7, column=1, value="ROUTEONE ECONTRACTS ")
    ws_summary.cell(row=8, column=1, value="  Reconciled (Matched + Amount Match)")
    ws_summary.cell(row=8, column=2, value=str(routeone_reconciled))
    ws_summary.cell(row=9, column=1, value="  Not Reconciled (In Vault, No Match)")
    ws_summary.cell(row=9, column=2, value=str(routeone_not_reconciled))
    ws_summary.cell(row=10, column=1, value="  Missing from Vault (In Funding Only)")
    ws_summary.cell(row=10, column=2, value=str(routeone_missing))
    ws_summary.cell(row=11, column=1, value="  Errors (Matched, Amount Mismatch)")
    ws_summary.cell(row=11, column=2, value=str(routeone_errors_count))
    ws_summary.cell(row=12, column=1, value="  RouteOne Econtracts Subtotal")
    ws_summary.cell(row=12, column=2, value=str(routeone_subtotal))
    
    ws_summary.cell(row=14, column=1, value="DT ECONTRACTS ")
    ws_summary.cell(row=15, column=1, value="  Reconciled (Matched + Amount Match)")
    ws_summary.cell(row=15, column=2, value=str(dt_reconciled))
    ws_summary.cell(row=16, column=1, value="  Not Reconciled (In Vault, No Match)")
    ws_summary.cell(row=16, column=2, value=str(dt_not_reconciled))
    ws_summary.cell(row=17, column=1, value="  Missing from Vault (In Funding Only)")
    ws_summary.cell(row=17, column=2, value=str(dt_missing))
    ws_summary.cell(row=18, column=1, value="  Errors (Matched, Amount Mismatch)")
    ws_summary.cell(row=18, column=2, value=str(dt_errors_count))
    ws_summary.cell(row=19, column=1, value="  DT Econtracts Subtotal")
    ws_summary.cell(row=19, column=2, value=str(dt_subtotal))
    
    ws_summary.cell(row=21, column=1, value=" PAPER CONTRACTS ")
    ws_summary.cell(row=22, column=1, value="  Paper Contracts (All Types)")
    ws_summary.cell(row=22, column=2, value=str(paper_total))
    
    ws_summary.cell(row=24, column=1, value="VERIFICATION ")
    ws_summary.cell(row=25, column=1, value="  Total Processed (Should Equal Funding Total)")
    ws_summary.cell(row=25, column=2, value=str(total_processed))
    ws_summary.cell(row=26, column=1, value="  Funding Total (Original)")
    ws_summary.cell(row=26, column=2, value=str(total_funding_records))
    
    # Set column widths
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 15

    print(f"[DEBUG] Summary sheet '{summary_sheet_name[:31]}' written with static values.")

    # Write each sheet manually
    for sheet_name, df in dfs_to_export.items():
        # Create new sheet
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet names limited to 31 chars
        
        # Convert DataFrame to list of lists for manual writing
        if not df.empty:
            # Check if this is a NOT RECONCILED sheet (has generic column names)
            if "NOT RECONCILED" in sheet_name and df.columns[0].startswith("Column_"):
                # Skip the generic column headers, just write the data
                df_str = df.astype(str)
                data = df_str.values.tolist()
            else:
                # Normal sheet - include column headers
                df_str = df.astype(str)
                data = [df_str.columns.tolist()] + df_str.values.tolist()
        else:
            # Empty DataFrame - just write headers
            data = [df.columns.tolist()] if not df.empty else [['No Data']]
        
        # Write data row by row, cell by cell
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_value))
                
                # Format header row (only for non-NOT RECONCILED sheets)
                if row_idx == 1 and "NOT RECONCILED" not in sheet_name:
                    cell.font = styles.Font(bold=True)
                    cell.border = styles.Border(
                        left=styles.Side(style='thin'),
                        right=styles.Side(style='thin'),
                        top=styles.Side(style='thin'),
                        bottom=styles.Side(style='thin')
                    )
        
        # Set column widths
        for col in range(1, len(data[0]) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    
    # Now apply specific formatting for certain sheets
    def format_specific_sheets(filepath):
        wb = load_workbook(filepath)
        
        for sheet in wb.worksheets:
            sheet_name = sheet.title
            
            # Special handling for Route One Vault sheet (reconciled)
            if "Route One Vault" in sheet_name and "NOT RECONCILED" not in sheet_name and "Errors" not in sheet_name:
                print(f"Formatting {sheet_name}...")
                
                # Get the data range
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Make header row bold with borders
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = styles.Font(bold=True)
                    cell.border = styles.Border(
                        left=styles.Side(style='thin'),
                        right=styles.Side(style='thin'),
                        top=styles.Side(style='thin'),
                        bottom=styles.Side(style='thin')
                    )
                
                # Highlight Paper contracts (check column Q for "Paper")
                for row in range(2, max_row + 1):  # Skip header row
                    if max_col >= 17:  # Make sure column Q exists
                        contract_type = sheet.cell(row=row, column=17).value  # Column Q (17th column)
                        if contract_type == "Paper":
                            for col in range(1, sheet.max_column + 1):
                                cell = sheet.cell(row=row, column=col)
                                cell.fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Special handling for Route One Vault NOT RECONCILED sheet
            elif "Route One Vault NOT RECONCILED" in sheet_name:
                print(f"Formatting {sheet_name}...")
                
                # Get the data range
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Track which rows are column headers (rows immediately after section headers)
                column_header_rows = []
                
                # First pass: identify section headers and column header rows
                for row in range(1, max_row + 1):
                    first_cell_value = str(sheet.cell(row=row, column=1).value)
                    
                    if first_cell_value in ["IN VAULT BUT NOT IN FUNDING", "IN FUNDING BUT NOT IN VAULT"]:
                        # The next row will be column headers
                        if row + 1 <= max_row:
                            column_header_rows.append(row + 1)
                
                # Second pass: format all rows
                for row in range(1, max_row + 1):
                    first_cell_value = str(sheet.cell(row=row, column=1).value)
                    
                    # Check if this is a section header
                    if first_cell_value in ["IN VAULT BUT NOT IN FUNDING", "IN FUNDING BUT NOT IN VAULT"]:
                        # Style section header row
                        for col in range(1, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            cell.font = styles.Font(bold=True, size=12)
                            cell.fill = styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell.border = styles.Border(
                                left=styles.Side(style='thin'),
                                right=styles.Side(style='thin'),
                                top=styles.Side(style='thin'),
                                bottom=styles.Side(style='thin')
                            )
                    elif row in column_header_rows:
                        # Style column headers
                        for col in range(1, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            cell.font = styles.Font(bold=True)
                            cell.border = styles.Border(
                                left=styles.Side(style='thin'),
                                right=styles.Side(style='thin'),
                                top=styles.Side(style='thin'),
                                bottom=styles.Side(style='thin')
                            )
                    else:
                        # Check for E Contract highlighting if column Q exists
                        if max_col >= 17:
                            contract_type = sheet.cell(row=row, column=17).value  # Column Q
                            if contract_type == "E Contract":
                                for col in range(1, max_col + 1):
                                    cell = sheet.cell(row=row, column=col)
                                    cell.fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Special handling for DT Vault NOT RECONCILED sheet
            elif "DT Vault NOT RECONCILED" in sheet_name:
                print(f"Formatting {sheet_name}...")
                
                # Get the data range
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Track which rows are column headers (rows immediately after section headers)
                column_header_rows = []
                
                # First pass: identify section headers and column header rows
                for row in range(1, max_row + 1):
                    first_cell_value = str(sheet.cell(row=row, column=1).value)
                    
                    if first_cell_value in ["IN VAULT BUT NOT IN FUNDING", "IN FUNDING BUT NOT IN VAULT"]:
                        # The next row will be column headers
                        if row + 1 <= max_row:
                            column_header_rows.append(row + 1)
                
                # Second pass: format all rows
                for row in range(1, max_row + 1):
                    first_cell_value = str(sheet.cell(row=row, column=1).value)
                    
                    # Check if this is a section header
                    if first_cell_value in ["IN VAULT BUT NOT IN FUNDING", "IN FUNDING BUT NOT IN VAULT"]:
                        # Style section header row
                        for col in range(1, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            cell.font = styles.Font(bold=True, size=12)
                            cell.fill = styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell.border = styles.Border(
                                left=styles.Side(style='thin'),
                                right=styles.Side(style='thin'),
                                top=styles.Side(style='thin'),
                                bottom=styles.Side(style='thin')
                            )
                    elif row in column_header_rows:
                        # Style column headers
                        for col in range(1, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            cell.font = styles.Font(bold=True)
                            cell.border = styles.Border(
                                left=styles.Side(style='thin'),
                                right=styles.Side(style='thin'),
                                top=styles.Side(style='thin'),
                                bottom=styles.Side(style='thin')
                            )
                    else:
                        # Check for E Contract highlighting if column Q exists
                        if max_col >= 17:
                            contract_type = sheet.cell(row=row, column=17).value  # Column Q
                            if contract_type == "E Contract":
                                for col in range(1, max_col + 1):
                                    cell = sheet.cell(row=row, column=col)
                                    cell.fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Special handling for Summary sheet
            elif "Summary" in sheet_name:
                print(f"Formatting {sheet_name}...")
                
                # Just make headers bold and set column widths
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = styles.Font(bold=True)
                
                # Set column widths
                sheet.column_dimensions['A'].width = 45
                sheet.column_dimensions['B'].width = 15
        
        wb.save(filepath)
        print(f"Formatting completed for: {filepath}")

    format_specific_sheets(output_path)
    print(f"Final Excel written and formatted: {output_path}")
   
    def archive_and_delete():
        # ARCHIVE: Copy input and output to Archive/month_year_timestamp folders


        archive_folder_name = f"{month_year.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        archive_base = ARCHIVE_PATH / archive_folder_name
        archive_input = archive_base / "Input"
        archive_output = archive_base / "Output"

        archive_input.mkdir(parents=True, exist_ok=True)
        archive_output.mkdir(parents=True, exist_ok=True)

        # Copy input files
        for path in matched_files.values():
            copy2(path, archive_input)

        # Copy output Excel file
        output_excel_file = output_path
        copy2(output_excel_file, archive_output)

        print(f"Archived input and output to: {archive_base}")

        # DELETE processed input files
        for path in matched_files.values():
            try:
                os.remove(path)
            except Exception as e:
                print(f"Could not delete {path}: {e}")
   
    archive_and_delete()


if __name__ == '__main__':
    print("Starting Dealer Track - Route One Reconciliation Report")
    main()
    print("Processed daily files")