"""
Output to Excel Module

This module handles Excel output generation with formatting,
including auto-fitting column widths using openpyxl.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from pathlib import Path

def format_excel_file(file_path: Path, sheet_name: str = 'Sheet1'):
    """
    Format an existing Excel file with auto-fit columns and header formatting.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the worksheet to format (default: 'Sheet1')
    """
    try:
        # Load the workbook
        workbook = load_workbook(str(file_path))
        
        if sheet_name not in workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in workbook")
            return
            
        worksheet = workbook[sheet_name]
        
        # Auto-fit column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row (bold and borders)
        if worksheet.max_row > 0:
            header_row = worksheet[1]
            bold_font = Font(bold=True)
            thin_border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
            
            for cell in header_row:
                cell.font = bold_font
                cell.border = thin_border
        
        # Save the formatted workbook
        workbook.save(str(file_path))
        print(f"Excel file formatted successfully: {file_path}")
        
    except Exception as e:
        print(f"Error formatting Excel file: {e}")

def write_and_format_excel(dataframe: pd.DataFrame, file_path: Path, sheet_name: str = 'Sheet1'):
    """
    Write DataFrame to Excel and apply formatting in one step.
    
    Args:
        dataframe: DataFrame to write to Excel
        file_path: Path where the Excel file should be saved
        sheet_name: Name of the worksheet (default: 'Sheet1')
    """
    try:
        # Write DataFrame to Excel
        dataframe.to_excel(file_path, sheet_name=sheet_name, index=False)
        
        # Apply formatting
        format_excel_file(file_path, sheet_name)
        
        print(f"Excel file created and formatted: {file_path}")
        
    except Exception as e:
        print(f"Error creating and formatting Excel file: {e}")
        raise
