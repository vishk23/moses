import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from pathlib import Path

def format_excel(
    file_path,
    currency_cols=None,
    percent_cols=None,
    date_cols=None,
    buffer=2
):
    """
    Format an Excel file in place using openpyxl.
    Args:
        file_path (str or Path): Path to the Excel file to format.
        currency_cols (list): List of column names to format as currency ($#,##0.00)
        percent_cols (list): List of column names to format as percent (0.00%)
        date_cols (list): List of column names to format as short date (mm/dd/yyyy)
        buffer (int): Extra characters to add to autofit width
    """
    file_path = Path(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col_idx = {col: idx+1 for idx, col in enumerate(header)}

    # Set formats
    if currency_cols:
        for col in currency_cols:
            if col in col_idx:
                for row in ws.iter_rows(min_row=2, min_col=col_idx[col], max_col=col_idx[col]):
                    for cell in row:
                        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    if percent_cols:
        for col in percent_cols:
            if col in col_idx:
                for row in ws.iter_rows(min_row=2, min_col=col_idx[col], max_col=col_idx[col]):
                    for cell in row:
                        cell.number_format = '0.00%'
    if date_cols:
        for col in date_cols:
            if col in col_idx:
                for row in ws.iter_rows(min_row=2, min_col=col_idx[col], max_col=col_idx[col]):
                    for cell in row:
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    # Autofit columns
    for idx, col in enumerate(header, 1):
        max_length = len(str(col))
        for cell in ws[get_column_letter(idx)]:
            if cell.row == 1:
                continue
            try:
                val = str(cell.value) if cell.value is not None else ''
            except:
                val = ''
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[get_column_letter(idx)].width = max_length + buffer

    wb.save(file_path)
