import requests
import pandas as pd
import json
import os
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import cdutils.distribution # type: ignore
import shutil

import win32com.client as win32 

import warnings

API_KEY = "ac15589a824557b5b4d2260b45438215"

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "output"

RATE_TABLE_TEMPLATE = [
    {"desc": "1 Yr Weekly Average Treasury", "sched": 4, "field": "RIFLGFCY01_N.WF", "source": "FRED", "key": "WGS1YR"},
    {"desc": "1 Yr Monthly Average Treasury", "sched": 5, "field": "RIFLGFCY01_N.M", "source": "FRED", "key": "GS1"},
    {"desc": "3 Yr Weekly Average Treasury", "sched": 6, "field": "RIFLGFCY03_N.WF", "source": "FRED", "key": "WGS3YR"},
    {"desc": "5 Yr Weekly Average Treasury", "sched": 17, "field": "RIFLGFCY05_N.WF", "source": "FRED", "key": "WGS5YR"},
    {"desc": "FHLBB 1 Yr Classic Rate\nCommercial FHLBB 1 Year Classic Regular", "sched": 22, "field": "1 YEAR REG.", "source": "FHLB", "key": "1Y Classic"},
    {"desc": "FHLBB 2 Yr Classic Rate\nCommercial FHLBB 2 Year Classic ADV", "sched": 20, "field": "2 YEAR REG.", "source": "FHLB", "key": "2Y Classic"},
    {"desc": "FHLBB 3 Yr Classic Rate\nCommercial FHLBB 3 Year Classic Regular", "sched": 131, "field": "3 YEAR REG.", "source": "FHLB", "key": "3Y Classic"},
    {"desc": "FHLBB 3 Yr CDA Rate\nCommercial FHLBB 3 Year Classic CDA", "sched": 225, "field": "3 YEAR CDA", "source": "FHLB", "key": "3Y CDA"},
    {"desc": "FHLBB 4 Yr Classic Rate", "sched": 175, "field": "4 YEAR REG.", "source": "FHLB", "key": "4Y Classic"},
    {"desc": "FHLBB 5 Yr Classic Rate\nCommercial FHLBB 5 Year Classic Regular", "sched": 18, "field": "5 YEAR REG.", "source": "FHLB", "key": "5Y Classic"},
    {"desc": "FHLBB 5 Yr CDA Rate\nCommercial FHLBB 5 Year Classic CDA", "sched": 110, "field": "5 YEAR CDA", "source": "FHLB", "key": "5Y CDA"},
    {"desc": "FHLBB 6 Yr Classic Rate", "sched": 176, "field": "6 YEAR REG.", "source": "FHLB", "key": "6Y Classic"},
    {"desc": "FHLBB 7 Yr Classic Rate\nCommercial FHLBB 7 Year Classic Regular", "sched": 21, "field": "7 YEAR REG.", "source": "FHLB", "key": "7Y Classic"},
    {"desc": "FHLBB 7 Yr CDA Rate\nCommercial FHLBB 7 Year Classic CDA", "sched": 19, "field": "7 YEAR CDA", "source": "FHLB", "key": "7Y CDA"},
    {"desc": "FHLBB 10 Yr Classic Rate\nCommercial FHLBB 10 Year Classic Reg", "sched": 23, "field": "10 YEAR REG.", "source": "FHLB", "key": "10Y Classic"},
    {"desc": "FHLBB 5 Yr Amortizing Advance Regular\nCommercial FHLBB 5 Yr/5 Yr Amortizing Advance Regular", "sched": 129, "field": "5-5 YEAR REG.", "source": "FHLB", "key": "5/5 Amortizing"},
    {"desc": "FHLBB 5 Yr/20 Yr Amortizing Advance Regular\nCommercial FHLBB 5 Yr/20 Yr Amortizing Advance Regular", "sched": 130, "field": "5-20 YEAR REG.", "source": "FHLB", "key": "5/20 Amortizing"},
    {"desc": "FHLBB 10 Yr Amortizing Advance Regular\nCommercial FHLBB 10 Yr Amortizing Advance Regular", "sched": 133, "field": "10-10 YEAR REG.", "source": "FHLB", "key": "10/10 Amortizing"},
    {"desc": "SOFR - 1 Month CME (Add to COCC and Laser Pro)", "sched": 230, "field": "1 Month CME Term", "source": "CME", "key": "1M"},
    {"desc": "SOFR - 30 Day Average (Add to COCC Only, not Laser Pro)", "sched": 236, "field": "30-DAY AVERAGE(%)", "source": "CME", "key": "30D"},
]


def fetch_full_fred_data(series_ids):
    fred_data = {}

    for series_id in series_ids:
        # Metadata call
        meta_url = "https://api.stlouisfed.org/fred/series"
        meta_params = {
            "series_id": series_id,
            "api_key": API_KEY,
            "file_type": "json"
        }

        title = ""
        last_updated = ""
        meta_response = requests.get(meta_url, params=meta_params)
        if meta_response.status_code == 200:
            meta_json = meta_response.json()
            seriess = meta_json.get("seriess", [])
            if seriess:
                title = seriess[0].get("title")
                last_updated = seriess[0].get("last_updated")

        # Observations call
        obs_url = "https://api.stlouisfed.org/fred/series/observations"
        obs_params = {
            "series_id": series_id,
            "api_key": API_KEY,
            "file_type": "json",
            "sort_order": "desc",
            "limit": 1
        }

        obs_response = requests.get(obs_url, params=obs_params)
        value, date = None, ""
        if obs_response.status_code == 200:
            obs_json = obs_response.json()
            observations = obs_json.get("observations", [])
            if observations:
                value = float(observations[0]["value"]) if observations[0]["value"] not in (".", "") else None
                date = observations[0]["date"]

        fred_data[series_id] = {
            "value": value,
            "date": date,
            "title": title,
            "last_updated": last_updated
        }

    return fred_data


def get_cme_term_sofr_data():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto("https://www.cmegroup.com/services/sofr-strip-rates", timeout=20000)
        page.wait_for_timeout(3000)

        body_text = page.evaluate("document.querySelector('body').innerText")
        data = json.loads(body_text)

        latest = data["resultsStrip"][0]
        return {
            "1M": float(next(r["price"] for r in latest["rates"]["sofrRatesFixing"] if r["term"] == "1M")),
            "30D": float(latest["average30day"]),
            "date": latest["date"]
        }


def fetch_fhlbb_rates():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        page.goto("https://www.fhlbboston.com/fhlbank-boston/rates#/long-term", timeout=20000)
        with page.expect_download() as download_info:
            page.click("text=Download Today's Rates – Excel")
        download = download_info.value

        file_path = Path(download.suggested_filename)
        download.save_as(str(file_path))

        df = pd.read_excel(file_path, header=None)
        file_path.unlink()
        assert(df.at[42, 1] == "1 YEAR")
        assert(df.at[31, 3] == "CDA")
        assert(df.at[71, 3] == "CDA")
        return {
            "Effective Date (Classics)": df.at[29, 1],
            "1Y Classic": df.at[42, 2],
            "2Y Classic": df.at[46, 2],
            "3Y Classic": df.at[50, 2],
            "3Y CDA": df.at[50, 3],
            "4Y Classic": df.at[52, 2],
            "5Y Classic": df.at[54, 2],
            "5Y CDA": df.at[54, 3],
            "6Y Classic": df.at[56, 2],
            "7Y Classic": df.at[58, 2],
            "7Y CDA": df.at[58, 3],
            "10Y Classic": df.at[64, 2],
            "Effective Date (Amortizing)": df.at[69, 1],
            "5/5 Amortizing": df.at[100, 2],
            "5/20 Amortizing": df.at[106, 2],
            "10/10 Amortizing": df.at[124, 2],
        }


def export_rate_table_to_excel(fred_data, cme_data, fhlb_data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from datetime import datetime

    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    # Timestamp
    now = datetime.now()
    timestamp_str = now.strftime("%B %d, %Y at %I:%M %p EST")
    filename_str = now.strftime("Rate_Report_%b_%d_%y_%H%M.xlsx")

    # Prepare rows
    rows = []
    for row in RATE_TABLE_TEMPLATE:
        source = row["source"]
        key = row["key"]
        rate, date = "", ""

        if source == "FRED":
            data = fred_data.get(key, {})
            rate = data.get("value")
            date = data.get("date")
        elif source == "CME":
            rate = cme_data.get(key)
            date = cme_data.get("date")
        elif source == "FHLB":
            rate = fhlb_data.get(key)
            date = fhlb_data.get("Effective Date (Classics)") or fhlb_data.get("Effective Date (Amortizing)")

        # Convert to float
        if isinstance(rate, str):
            try:
                rate = float(rate.replace(",", ""))
            except:
                pass

        rows.append({
            "Index Description / Laser Pro Description": row["desc"],
            "Schedule\nNumber": row["sched"],
            "New Rate": rate,
            "Observation Date": date,
            "Field Name": row["field"]
        })

    df = pd.DataFrame(rows)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rate Table"

    # Styles
    header_font = Font(bold=True, size=13)
    body_font = Font(size=12)
    rate_font = Font(size=15, bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Freeze header row (row 2)
    ws.freeze_panes = "A3"

    # Write report info merged across row 1
    info_text = "Reviewed By:__________________________________________"
    ws.merge_cells("A1:E1")
    info_cell = ws["A1"]
    info_cell.value = info_text
    info_cell.font = Font(size=18)
    info_cell.alignment = Alignment(wrap_text=False, vertical="top")
    
    info_text = f"Generated on {timestamp_str} by BCSB Data and Analytics\nBusinessintelligence@bcsbmail.com"
    ws.merge_cells("A2:E2")
    info_cell = ws["A2"]
    info_cell.value = info_text
    info_cell.font = Font(italic=True, size=11)
    info_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        ws.append(row)

        if r_idx == 3:
            for cell in ws[r_idx]:
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
        else:
            ws.row_dimensions[r_idx].height = 45
            if r_idx % 2 == 0:
                for cell in ws[r_idx]:
                    cell.fill = zebra_fill
            for c_idx, cell in enumerate(ws[r_idx], start=1):
                cell.font = body_font
                cell.border = border
                if c_idx == 1:
                    cell.alignment = wrap_align
                elif c_idx == 3 and isinstance(cell.value, (int, float)):
                    cell.number_format = '#0.00#############'

                    cell.font = rate_font
                    cell.alignment = center_align
                else:
                    cell.alignment = center_align

    # Adjust column widths
    column_widths = {
        1: 42,
        2: 12,
        3: 18,
        4: 20,
        5: 25
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width
    


    output_path = OUTPUT_DIR / filename_str
    pdf_path = OUTPUT_DIR  / filename_str
    pdf_path = str(pdf_path).replace(".xlsxpdf","")
    pdf_path = pdf_path.replace(".xlsx","")

    # Save
    wb.save(output_path)
    print(f"Excel table written to {output_path}")
    
    win32.gencache.is_readonly = False
    win32.gencache.Rebuild()
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(output_path)
    ws = wb.Worksheets(1)
    ws.PageSetup.Orientation = 1
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.TopMargin = 0.5
    ws.PageSetup.BottomMargin = 0.2
    wb.ExportAsFixedFormat(0, str(pdf_path))
    wb.Close(False)
    excel.Quit()
    print(f"Excel table written to PDF {pdf_path}")

    # Distribution
    recipients = [
        "Kelly.Abernathy@bcsbmail.com",
        "Taylor.Willbanks@bcsbmail.com",
        "Josephine.Willard@bcsbmail.com",
        "Patty.DeSimone@bcsbmail.com",
        "Zachary.Cabral@bcsbmail.com",
        "Tanner.Vickery@bcsbmail.com"
    ]
    cc_recipients = None
    bcc_recipients = [
        "chad.doorley@bcsbmail.com",
        "businessintelligence@bcsbmail.com",
    ]

    time_subjectline = now.strftime("%B %d, %Y")

    subject = f"Daily Rate Scrape - {time_subjectline}" 
    body = "Hi, \n\nAttached is the Daily Rate Scrape Report. If you have any questions, please reach out to BusinessIntelligence@bcsbmail.com \n\nThanks!"
    PDF_ATTACHMENT = Path(pdf_path + '.pdf')  
    # print(PDF_ATTACHMENT)
    attachment_paths = [PDF_ATTACHMENT]
    cdutils.distribution.email_out(recipients, cc_recipients, bcc_recipients, subject, body, attachment_paths)

    # Drop in Ops folder
    OPS_OUTPUT_PATH = Path(r"\\00-berlin\Operations\Loan Servicing\Daily Rate Updates\Output")
    OPS_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(PDF_ATTACHMENT, OPS_OUTPUT_PATH)




if __name__ == '__main__':
    
    print("Running Rate Scraper")
    fred_series_ids = ["WGS1YR", "GS1", "WGS3YR", "WGS5YR"]
    fred_data = fetch_full_fred_data(fred_series_ids)
    cme_data = get_cme_term_sofr_data()
    fhlb_data = fetch_fhlbb_rates()
    export_rate_table_to_excel(fred_data, cme_data, fhlb_data)
    print("Finished Rate Scraper")