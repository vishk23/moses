"""
Main Entry Point
"""
import pandas as pd
import numpy as np

from pathlib import Path
from openpyxl import load_workbook
from deltalake import DeltaTable

import src.config
import os
import shutil
import cdutils.distribution # type: ignore
from src._version import __version__
from src.config import BASE_PATH, INPUT_DIR, OUTPUT_DIR, EMAIL_TO, EMAIL_CC, EMAIL_BCC, EXCEPTION_EMAIL_TO, EXCEPTION_EMAIL_CC, EXCEPTION_EMAIL_BCC
from src.daily_mismatch_txns.api_call import (
    fetch_latest_to_input,
)


def main():
    ASSETS_PATH = BASE_PATH / Path('./assets')
    ASSETS_PATH.mkdir(parents=True, exist_ok=True)
    # Fetch latest .prn/.txt into INPUT_DIR so downstream reads it
    try:
        dest = fetch_latest_to_input("CO_VSUS", storage_type_id=1, filename_template="CO_VSUS_{pkid}.txt")
        if dest:
            print(f"Downloaded and saved latest CO_VSUS file to INPUT_DIR: {dest}")
        else:
            print("No CO_VSUS documents found via API search.")
    except Exception as e:
        print(f"Warning: API fetch skipped/failed: {e}")
    
    # Choose newest .txt file and archive the rest
    txt_files = sorted(INPUT_DIR.glob("*.txt"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not txt_files:
        raise FileNotFoundError(f"No .txt files found in {INPUT_DIR}")
    input_src_path = txt_files[0]
    # Archive any extras to keep the folder clean for next runs
    for extra in txt_files[1:]:
        shutil.move(extra, INPUT_DIR / "archive" / extra.name)
        print(f"Archived extra file: {extra.name}")
    file_to_move = input_src_path.name


    column_names = [
        "Card Nbr",
        "Acct Nbr",
        "Trans Amt",
        "RTXN Typ",
        "RetRefNbr",
        "Comment",
        "Merchant"
    ]

    column_widths = [16, 20, 18, 18, 12, 35, 40]

    df = pd.read_fwf(
        input_src_path,
        widths=column_widths,
        names=column_names,
        encoding='latin1'
    )

    df['Card and Acct and Tran Amt'] = df['Card Nbr'] + df['Acct Nbr'] + df['Trans Amt']

    # locating row index that starts Records On Fiserv ACH File but not ON AT_CACT File which separates top half of the report from bottom half
    idx = df[df['RetRefNbr'].str.contains('t not ON AT_', case=False, na=False)].index[0]

    top = df.loc[:idx]
    top = top[top['Card Nbr'].str.contains('5', na=False)]
    bottom = df.loc[idx:]
    bottom = bottom[bottom['Card Nbr'].str.contains('5', na=False)]

    # now we just find unique occurances of Card and Acct and Tran Amt (occurances from either dataframe that isn't in the other one)
    unique_in_top = top[~top['Card and Acct and Tran Amt'].isin(bottom['Card and Acct and Tran Amt'])]
    unique_in_bottom = bottom[~bottom['Card and Acct and Tran Amt'].isin(top['Card and Acct and Tran Amt'])]
    result = pd.concat([unique_in_top, unique_in_bottom])
    result = result[['Card Nbr', 'Acct Nbr', 'Trans Amt', 'RTXN Typ', 'RetRefNbr', 'Merchant']]

    field_widths = [120, 10]
    field_names = ['Not Needed', 'Date']
    input2 = pd.read_fwf(
        input_src_path,
        widths=field_widths,
        names=field_names,
        encoding='latin1'
    )
    input2['RecordID'] = pd.Series(range(1, len(input2) + 1), dtype='int32')
    
    # adding date to result dataframe
    date_str = str(input2.at[1, 'Date']).strip()
    result['Date'] = date_str
    result['Source'] = "co_vsus re-run"
    
    result['Trans Amt'] = result['Trans Amt'].str.replace(",","")
    result['Trans Amt'] = pd.to_numeric(result['Trans Amt'], errors="coerce")
    result['Trans Amt'] = np.where(result['RTXN Typ'] == "PWTH", 0 - result['Trans Amt'], result['Trans Amt'])
    result = result[['Card Nbr', 'Date', 'Source', 'Trans Amt', 'Acct Nbr', 'RTXN Typ', 'RetRefNbr', 'Merchant']]


    cardnbrs = list(result['Card Nbr'])
    acctnbrs = list(result['Acct Nbr'])
    amount = list(result['Trans Amt'])
    deb_or_cred = []
    merchants = list(result['Merchant'])

    # Save unadjusted version of account numbers because we pad with leading zeros
    # Necsesary for later merge against COCC active accounts
    unadjusted_card = cardnbrs.copy()
    unadjusted_acctnbr = acctnbrs.copy()

    for i in range(len(result)):

        # left padding acctnbrs with 0s
        while len(acctnbrs[i]) < 12:
            acctnbrs[i] = '0' + acctnbrs[i]

        # determining if debit or credit based on sign of Trans Amt
        if amount[i] < 0:
            deb_or_cred.append("Debit")
            amount[i] = amount[i] * -1  # making all amounts positive
        elif amount[i] > 0:
            deb_or_cred.append("Credit")
        else:
            pass
        
        # appending last 4 digits of card number to the merchant

        merchants[i] += f" ({cardnbrs[i][-4:]})"

    # Matching against COCC
    temp_df = pd.DataFrame({
        'cardnbr': unadjusted_card,
        'acctnbr': unadjusted_acctnbr,
        'amount': amount,
        'merchant': merchants
    }).copy()

    # Make string if not already
    temp_df['acctnbr'] = temp_df['acctnbr'].astype(str)

    
    active_accts = DeltaTable(src.config.SILVER / "account").to_pandas()
    
    # Merging
    merged_df = pd.merge(temp_df, active_accts, how='outer', on='acctnbr', indicator=True)

    # Exceptions dataframe creation
    exceptions = merged_df[merged_df['_merge'] == 'left_only'].copy()

    # # Write out exceptions here to check if this works
    # TEMP_OUTPUT = Path(r"C:\Users\w322800\Documents\gh\bcsb-prod\Reports\Risk Management\Daily Mismatched Debit Card Txns\bin\exceptions.parquet")
    # exceptions.to_parquet(TEMP_OUTPUT, index=False)

    # Save for separate distribution
    exceptions = exceptions[[
        'cardnbr',
        'acctnbr',
        'amount',
        'merchant'
    ]].copy()

    # move txt file to archive
    input_archive_path = INPUT_DIR / Path('./archive') / Path(file_to_move)
    shutil.move(input_src_path, input_archive_path)
    print(f"Moved {file_to_move} to input/archive directory.")

    # filling in template
    wb = load_workbook(ASSETS_PATH / Path("txtparser_template.xlsx"))
    ws = wb.active

    for i, value in enumerate(deb_or_cred):
        ws.cell(row=8+2*i, column=1, value=value)
    for i, value in enumerate(acctnbrs):
        ws.cell(row=8+2*i, column=3, value=value)
    for i, value in enumerate(amount):
        ws.cell(row=8+2*i, column=5, value=value)
    for i, value in enumerate(merchants):
        ws.cell(row=8+2*i, column=7, value=value)


    os.makedirs(OUTPUT_DIR, exist_ok=True)
    # output_date_str = f"{today.month}.{today.day:02}.{today.year % 100:02}"
    # date_str = f"{today.month}/{today.day:02}/{today.year % 100:02}"
    filename = "Daily Posting Sheet " + date_str + ".xlsx"
    output_file = os.path.join(OUTPUT_DIR,filename)

    ws['C3'] = f"{date_str}"

    # before saving, move everything in output folder to output/archive
    for file in OUTPUT_DIR.glob("*.xlsx"):
        file_to_move = file.name
        src_path = OUTPUT_DIR / Path(file_to_move)
        output_archive_path = OUTPUT_DIR / Path('./archive') / Path(file_to_move)
        shutil.move(src_path, output_archive_path)
        print(f"Moved {file_to_move} to output/archive directory.")

    # save spreadsheet to output directory
    wb.save(output_file)
    print(f"Report saved to {output_file}")

    # Usage
    # # Distribution
    subject = f"Daily Transaction Mismatch for Posting - {date_str}" 
    body = "Hi all, \n\nPlease see the attached Daily Transaction Mismatch file for Posting." \
    "\n\nPlease reach to Patrick Quinn (patrick.quinn@bcsbmail.com) or BusinessIntelligence@bcsbmail.com if you have any questions or issues."
    attachment_paths = [output_file]

    # Send main file to deposit ops
    cdutils.distribution.email_out(EMAIL_TO, EMAIL_CC, EMAIL_BCC, subject, body, attachment_paths)

    EXCEPTION_FILENAME = "Exceptions " + date_str + ".txt"
    EXCEPTION_OUTPUT = OUTPUT_DIR / EXCEPTION_FILENAME

    # Define the email subject and body, which we may adjust if there are no exceptions.
    subject = f"Exceptions: Daily Transaction Mismatch for Posting - {date_str}"
    body = (
        "Hi all,\n\n"
        "Please see the attached exceptions regarding the Daily Transaction Mismatch file for Posting.\n\n"
        "This shows the items that have account numbers that don't match active accounts in COCC.\n\n"
        "Please reach to BusinessIntelligence@bcsbmail.com if you have any questions or issues."
    )

    if exceptions.empty:
        # If the exceptions DataFrame is empty, create a simple message file and update the email body.
        report_content = "No transaction mismatch exceptions found for this reporting period."
        body = "Hi all,\n\nThere were no transaction mismatch exceptions found for today's posting file."
        print("No exceptions found. A note will be written to the output file.")

    else:
        # If there are exceptions, build the formatted fixed-width report.
        df = exceptions.copy() # Work on a copy to avoid SettingWithCopyWarning
        
        # Convert all columns to string to measure length accurately
        df['cardnbr'] = df['cardnbr'].astype(str)
        df['acctnbr'] = df['acctnbr'].astype(str)
        # Ensure amount has two decimal places for consistent formatting
        df['amount'] = df['amount'].apply(lambda x: f"{x:.2f}")

        # --- Calculate column widths ---
        # Start with header length, then find the max length in each column.
        # Add padding for space between columns.
        padding = 2
        col_widths = {
            col: max(df[col].str.len().max(), len(col)) + padding
            for col in df.columns
        }

        # --- Build the report string line-by-line ---
        report_lines = []
        
        # 1. Create the main header for the report
        title = " Daily Transaction Mismatch Exceptions Report "
        total_width = sum(col_widths.values())
        report_lines.append("=" * total_width)
        report_lines.append(f"=={title.center(total_width - 4)}==")
        report_lines.append("=" * total_width)
        report_lines.append("") # Blank line
        
        # 2. Add metadata
        report_lines.append(f"Report Date: {date_str}")
        report_lines.append(f"Total Exceptions: {len(df)}")
        report_lines.append("") # Blank line

        # 3. Create the column headers row
        header_line = "".join([col.ljust(col_widths[col]) for col in df.columns])
        report_lines.append(header_line)

        # 4. Create the separator line under the headers
        separator_line = "".join(("-" * (col_widths[col] - padding)).ljust(col_widths[col]) for col in df.columns)
        report_lines.append(separator_line)

        # 5. Add each data row, formatted to the correct width
        for index, row in df.iterrows():
            data_line = "".join([str(row[col]).ljust(col_widths[col]) for col in df.columns])
            report_lines.append(data_line)
        
        # 6. Add a footer
        report_lines.append("") # Blank line
        report_lines.append("--- End of Report ---")

        # Join all lines into a single string for writing
        report_content = "\n".join(report_lines)

    # --- Write the generated content to the .txt file ---
    with open(EXCEPTION_OUTPUT, "w", encoding="utf-8") as f:
        f.write(report_content)
    print(f"Formatted text exception report saved to {EXCEPTION_OUTPUT}")


    # # Distribution (This section now uses the conditionally-set body)
    attachment_paths = [output_file, EXCEPTION_OUTPUT]

    # Send exceptions email
    cdutils.distribution.email_out(EXCEPTION_EMAIL_TO, EXCEPTION_EMAIL_CC, EXCEPTION_EMAIL_BCC, subject, body, attachment_paths)

if __name__ == '__main__':
    print(f"Starting [{__version__}]")
    main()
    print("Complete!")

