# Environment
import pandas as pd
from pandas.tseries.offsets import MonthEnd
import time
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, NamedStyle
from openpyxl.styles.numbers import NumberFormat
import os

def export_and_format(current_report, current_report_raw, new_additions, removed, repo, old_sum):

    start_time = time.time()


    #################################
    # Most recent month end &amp; prior month end
    today = datetime.today()
    if today.day == 1:
        most_recent_month_end = today - timedelta(days=1)
    else:
        most_recent_month_end = today.replace(day=1) - timedelta(days=1)
    prior_month_end =  most_recent_month_end - MonthEnd(1)
    most_recent_month_end_str = most_recent_month_end.strftime("%m/%d/%y")
    prior_month_end_str = prior_month_end.strftime("%m/%d/%y")


    #################################
    blank_df = {
        'TagType':['Blank']
    }
    blank_df = pd.DataFrame(blank_df)


    #################################
    dfs_to_union = [new_additions, blank_df, removed, blank_df, repo]

    reconciliation_df = pd.concat(dfs_to_union, ignore_index=True)


    #################################
    current_sum = current_report_raw['Net Balance'].sum()
    net_adjustments = current_sum - old_sum['Net Balance'].iloc[0]

    #################################
    recon_sum = reconciliation_df['Total Change'].sum()


    #################################
    principal_payments = net_adjustments - recon_sum
    # principal_payments


    #################################
    output_dir = r'\\00-DA1\Home\Share\\Data & Analytics Initiatives\\Project Management\\Credit_Loan_Review\\Resolution Committee Automation\\Non Accruals\\Production\\Output'
    os.makedirs(output_dir, exist_ok=True)
    last_month = date.today().replace(day=1) - timedelta(1)  
    date_str = last_month.strftime("%B %Y")
    filename = "NonAccruals " + date_str + ".xlsx"
    output_file = os.path.join(output_dir,filename)

    print(output_file)

    #################################
    columns_to_drop = ['MJACCTTYPCD','CURRMIACCTTYPCD','CURRACCTSTATCD','EFFDATE','NEXTDUEDATE','TOTALPI','Current Balance','Charged Off','FDIC Code','Non Accrual Date','Report']
    reconciliation_df = reconciliation_df.drop(columns=[col for col in columns_to_drop if col in reconciliation_df.columns])


    #################################
    # Making everything into a dataframe
    old_me_df = {
        'Col1':['',''],
        'Col2':['',''],
        'Customer Name':['Reconciliation', prior_month_end_str],
        'Col4':['',''],
        'Col5':['',''],
        'Col6':['',''],
        'Total Change': ['', old_sum['Net Balance'].iloc[0]],
        'Count':['', old_sum['Count'].iloc[0]]
    }

    old_me_df = pd.DataFrame(old_me_df)

    summary_df = {
        'Col1':['','',''],
        'Col2':['','',''],
        'Customer Name':['PRINCIPAL PAYMENTS', 'TOTAL ADJUSTMENTS', most_recent_month_end_str],
        'Col4':['','',''],
        'Col5':['','',''],
        'Col6':['','',''],
        'Total Change': [principal_payments, net_adjustments, current_sum],
        'Count':['', '', len(current_report_raw)],
        'Col9':['','',''],
        'Col10':['','',''],
        'Col11':['','',''],
        'Col12':['','',''],
        'TagType':['','','Upper_Sum'] # This is a formatting tag for the new total and count
    }

    summary_df = pd.DataFrame(summary_df)

    #################################
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:

        # 4 - CML Reconciliation
        old_me_df.to_excel(writer, index=False, sheet_name='Sheet1', startcol=0, startrow=5, header=False)
        
        start_row = 5 + len(old_me_df) + 1
        
        reconciliation_df.to_excel(writer, index=False, sheet_name='Sheet1', startcol=0, startrow=start_row, header=True)
        
        start_row = start_row + len(reconciliation_df) + 2
        start_save = start_row
        
        summary_df.to_excel(writer, index=False, sheet_name='Sheet1', startcol=0, startrow=start_row, header=False)
        
        start_row = start_row + len(summary_df) + 2
        end_save = start_row
        
        current_report.to_excel(writer, index=False, sheet_name='Sheet1', startcol=0, startrow=start_row, header=True)
        
        
        
    wb = load_workbook(output_file)
    ws = wb['Sheet1']

    # Dimensions
    ws.column_dimensions['A'].width = 28.33
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 17
    ws.column_dimensions['L'].width = 13


    # Cell Font &amp; Formatting
    upper_section_font = Font(size=10, name='Arial', bold=True)
    bold_italics_font = Font(size=10, name='Arial', bold=True, italic=True)
    subtitle_font = Font(size=10, name='Arial', bold=True)
    data_font = Font(size=10, name='Arial')
    wrap_alignment = Alignment(wrap_text=True, vertical='bottom', horizontal='center')
    center_alignment = Alignment(horizontal='center')
    title_block = Font(size=14, name='Arial', bold=True)
    accounting_style = '"$"#,##0.00_);("$"#,##0.00)'
    count_format = '#,##0_);(#,##0)'

    # Borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thin'))
    double_bottom_border= Border(bottom=Side(style='double'))
    side_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    top_border = Border(top=Side(style='thin'))


    # Header
    ws['A1'] = "BRISTOL COUNTY SAVINGS BANK"
    ws['A1'].font = upper_section_font
    ws['A2'] = "MONTHLY NON ACCRUAL LOAN LIST"
    ws['A2'].font = upper_section_font
    ws['A3'] = f"{most_recent_month_end_str}"
    ws['A3'].font = upper_section_font

    last_row = ws.max_row

    # Set column B to bold                
    for row in range(1, 8):
        cell = ws.cell(row=row, column=3)
        cell.font = upper_section_font
        
    for row in range(start_save, end_save):
        cell = ws.cell(row=row, column=3)
        cell.font = upper_section_font

    for row in range(1, last_row+1):
        cell = ws[f'E{row}']
        cell.alignment = center_alignment
        
    # Check column M for 'TagType' and assign Wrap Text and centering
    for row in ws.iter_rows(min_col=13, max_col=13):
        for cell in row:
            row_idx = cell.row
            if cell.value == 'TagType':
            
                for col in ws.iter_cols(min_row=row_idx, max_row=row_idx):
                    for cell in col:
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            
            elif cell.value == 'Sum':
                # Bold the cells in the subtotal rows and add borders
                for col in ws.iter_cols(min_col=6, max_col=8, min_row=row_idx, max_row=row_idx):
                    for cell in col:
                        cell.font = bold_italics_font
                        cell.border = top_border + double_bottom_border
                ws.cell(row=row_idx, column=10).font = bold_italics_font
                
            elif cell.value == 'Title':
                for col in ws.iter_cols(min_row=row_idx, max_row=row_idx):
                    for cell in col:
                        cell.font = upper_section_font

                
    # Apply Number formatting &amp; Alignment
    for row in range(1, start_row+1):
        for col in ['F','G']:
            cell = ws[f'{col}{row}']
            cell.number_format = accounting_style
        cell_h = ws[f'H{row}']
        cell_h.number_format = count_format


    for row in range(start_row+1, last_row+1):
        for col in ['F','G','H']:
            cell = ws[f'{col}{row}']
            cell.number_format = accounting_style
            
    for row in range(1, last_row+1):
        cell = ws[f'L{row}']
        cell.number_format = 'mm/dd/yyyy'
        
    for row in range(1, last_row+1):
        cell = ws[f'J{row}']
        cell.alignment = center_alignment
        
    for row in range(1, start_row+1):
        cell = ws[f'H{row}']
        cell.alignment = center_alignment
        
    for row in range(1, last_row+1):
        cell = ws[f'I{row}']
        cell.alignment = center_alignment
        
    for row in range(1, last_row+1):
        cell = ws[f'K{row}']
        cell.number_format = accounting_style
        
    # Clearing Out Columns M &amp; N
    for row in range(1, last_row+1):
        for col in [13, 14]:
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

                
    # Saving
    wb.save(output_file)

    print(f"Report saved to {output_file}")

    # A start time variable was created at the beginning to track the number of seconds this script takes to execute.
    print(f"Script took {time.time() - start_time} seconds.")