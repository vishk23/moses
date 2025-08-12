from sqlalchemy import text
import sys
import os
import cdutils.database.connect  # type: ignore

import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment

import numpy as np
from pathlib import Path
from decimal import Decimal


import src.config

"""
Financial Difficulty Modifications Report - Main Entry Point

BUSINESS LOGIC EXTRACTED:

Data Sources & Relationships:
- COCCDM.WH_ACCTCOMMON: Account master data with status, balances, and product information
- COCCDM.WH_ACCTLOAN: Risk ratings, charged-off balances, and current due dates
- OSIBANK.WH_ACCTUSERFIELDS: TDR (Troubled Debt Restructuring) flags and dates
- Uses business month-ends for database queries and actual month-ends for DPD calculations

Business Rules:
- Only includes loans with TDR flag = 'Y' and valid TDR date on or before report date
- Net Balance = Note Balance - Charged Off Amount
- Days Past Due calculated from actual month-end to current due date (minimum 0)
- Major account type override: CM15/CM16 → Consumer (CNS) instead of Commercial (CML)
- Accounts with zero net balance have DPD reset to 0

Data Processing Flow:
1. Fetch current and prior month snapshots using business month-ends
2. Merge account, loan, and user field data for TDR identification
3. Filter to active/non-performing accounts with valid TDR flags and dates
4. Calculate Days Past Due using actual month-end dates for accuracy
5. Perform reconciliation analysis (new additions, removals, advances, paydowns)
6. Generate delinquency summary buckets (0-29, 30-59, 60+ days) by major type
7. Export formatted Excel with reconciliation, summary, and detailed account listing

Key Calculations:
- Net Balance = Note Balance - Charged Off Amount
- Days Past Due = Actual Month End - Current Due Date (clipped at 0)
- TDR Date validation against report effective date
- Reconciliation categories: additions, closures, charge-offs, LOC advances/paydowns

Business Intelligence Value:
Tracks troubled debt restructuring modifications for regulatory compliance and Resolution
Committee oversight, providing detailed reconciliation and delinquency analysis of modified
loans with comprehensive month-over-month change tracking.
"""


def get_last_2_month_ends():
    """Get the last 2 business month ends (for database queries)"""
    ends = pd.date_range(end=pd.Timestamp.today(), periods=2, freq="BME")
    return [d.to_pydatetime().date() for d in ends]


def get_actual_month_ends():
    """Get the last 2 actual month ends (for DPD calculations)"""
    ends = pd.date_range(end=pd.Timestamp.today(), periods=2, freq="ME")  # Month End, not Business Month End
    return [d.to_pydatetime().date() for d in ends]


# ----------------------------------
# DATA FETCHING FUNCTIONS
# ----------------------------------
def fetch_snapshot_common(month_end: str) -> pd.DataFrame:
    """
    Fetch the WH_ACCTCOMMON snapshot for a given month-end date.
    """
    effdate_sql = f"TO_DATE('{month_end}', 'YYYY-MM-DD')"
    
    sql = text(f"""
    SELECT
        ACCTNBR,
        MJACCTTYPCD,
        CURRMIACCTTYPCD,
        PRODUCT,
        CURRACCTSTATCD,
        OWNERNAME,
        NOTEBAL,
        EFFDATE
    FROM COCCDM.WH_ACCTCOMMON
    WHERE EFFDATE = {effdate_sql}
    """)
    
    queries = [{'key': 'wh_acctcommon', 'sql': sql, 'engine': 2}]
    df = cdutils.database.connect.retrieve_data(queries)
    return df


def fetch_snapshot_loan(month_end: str) -> pd.DataFrame:
    """
    Fetch the WH_ACCTLOAN snapshot for a given month-end date.
    """
    effdate_sql = f"TO_DATE('{month_end}', 'YYYY-MM-DD')"
    
    sql = text(f"""
    SELECT
        ACCTNBR,
        CURRDUEDATE,
        RISKRATINGCD AS Risk,
        COBAL
    FROM COCCDM.WH_ACCTLOAN
    WHERE EFFDATE = {effdate_sql}
    """)
    
    queries = [{'key': 'wh_acctloan', 'sql': sql, 'engine': 2}]
    return cdutils.database.connect.retrieve_data(queries)


def fetch_user_fields(month_end: str) -> pd.DataFrame:
    """
    Fetch account user fields (TDR date, TDR note, etc.) for a given month-end.
    """
    sql = text(f"""
    SELECT
        ACCTNBR,
        ACCTUSERFIELDCD,
        ACCTUSERFIELDVALUE
    FROM OSIBANK.WH_ACCTUSERFIELDS
    WHERE ACCTUSERFIELDCD IN ('TDAT','TDR','TNOT')
    """)
    
    queries = [{'key': 'wh_acctuserfields', 'sql': sql, 'engine': 1}]
    df = cdutils.database.connect.retrieve_data(queries)
    df = df.get("wh_acctuserfields")
    
    return (
        df
        .pivot_table(index='acctnbr', columns='acctuserfieldcd', values='acctuserfieldvalue', aggfunc='first')
        .reset_index()
    )


def fetch_data_for_date(snapshot_date: str, month_end_date: str) -> pd.DataFrame:
    """
    Build the 'current_report' DataFrame for a single month-end by
    merging common, loan, and user-fields snapshots and applying
    the business logic transformations.
    
    Args:
        snapshot_date: Business day date to query from database (YYYY-MM-DD format)
        month_end_date: Actual month-end date for DPD calculations (YYYY-MM-DD format)
    """
    common = fetch_snapshot_common(snapshot_date).get('wh_acctcommon')
    loan = fetch_snapshot_loan(snapshot_date).get("wh_acctloan")
    user = fetch_user_fields(snapshot_date)
    
    # Override Major based on CURRMIACCTTYPCD
    common['Major'] = np.where(
        common['currmiaccttypcd'].isin(['CM15','CM16']),
        'CNS',
        common['mjaccttypcd']
    )

    # Subtotal logic
    common['Product Name'] = common['product']
    common.loc[(common['Major']=='MTG') & common['product'].isna(), 'Product Name'] = 'Subtotal: Mortgage'
    common.loc[(common['Major']=='CNS') & common['product'].isna(), 'Product Name'] = 'Subtotal: Consumer'
    common.loc[(common['Major']=='CML') & common['product'].isna(), 'Product Name'] = 'Subtotal: Commercial'

    # Merge loan and user-fields
    df = common.merge(
        loan[['acctnbr','currduedate','risk','cobal']],
        on='acctnbr', how='left'
    ).merge(
        user[["acctnbr","TDAT","TDR","TNOT"]], on='acctnbr', how='left'
    )
    
    # Rename columns
    df = df.rename(columns={
        'notebal': 'Note Balance',
        'cobal': 'Charged Off',
        'ownername': 'Customer Name',
        'acctnbr': 'Account Number'
    })

    # Calculate Net Balance and Days Past Due
    df['Net Balance'] = df['Note Balance'] - df['Charged Off']
    
    # Use the actual month-end date for DPD calculation, not the snapshot date
    month_end_date_dt = pd.to_datetime(month_end_date)
    df['Days Past Due'] = (
        month_end_date_dt - pd.to_datetime(df['currduedate'])
    ).dt.days.clip(lower=0)
    
    # Filter active/non-performing and TDR entries
    df = df[df['curracctstatcd'].isin(['ACT','NPFM'])]

    # Only include loans where TDR flag is 'Y' and TDR date is on or before the report's effective date
    if 'TDR' in df.columns and 'TDAT' in df.columns:
        df['TDAT'] = pd.to_datetime(df['TDAT'], errors='coerce')
        report_date = pd.to_datetime(month_end_date)
        df = df[(df['TDR'] == 'Y') & (df['TDAT'].notna()) & (df['TDAT'] <= report_date)]

    return df


def retrieve_data():
    """
    Retrieve all needed DataFrames by fetching current- and prior-month snapshots.
    """
    business_month_ends = get_last_2_month_ends()
    prior_business_end = business_month_ends[0]
    current_business_end = business_month_ends[1]
    
    actual_month_ends = get_actual_month_ends()
    prior_actual_end = actual_month_ends[0]
    current_actual_end = actual_month_ends[1]
    
    current_report = fetch_data_for_date(str(current_business_end), str(current_actual_end))
    prior_report = fetch_data_for_date(str(prior_business_end), str(prior_actual_end))

    acctcommon = fetch_snapshot_common(str(current_business_end)).get("wh_acctcommon")[['acctnbr','curracctstatcd']]

    prior_date_df = pd.DataFrame({'Prior Month End': [str(prior_actual_end)]})

    return {
        'current_report': current_report,
        'acctcommon': acctcommon,
        'prior_report': prior_report,
        'prior_date_df': prior_date_df
    }



def main():
    """Main report execution function"""
    print("Starting Resolution Committee Financial Difficulty Modifications Report")
    
    actual_month_ends = get_actual_month_ends()
    prev_end = str(actual_month_ends[0])
    curr_end = str(actual_month_ends[1])
    
    curr_end_xlsx = curr_end + "_FDM_report.xlsx"
    
    output_path = src.config.OUTPUT_DIR / curr_end_xlsx
    
    data = retrieve_data()
    
    current_date = actual_month_ends[1].strftime("%m/%d/%y")
    prior_date = actual_month_ends[0].strftime("%m/%d/%y")
    
    reconciliation_full, summary_combined_df, current_report, current_date, prior_date = report_creation(
        data['current_report'], 
        data['acctcommon'], 
        data['prior_report'], 
        data['prior_date_df'],
        current_date,
        prior_date
    )
    
    output_to_excel(str(src.config.OUTPUT_DIR), curr_end_xlsx, reconciliation_full, summary_combined_df, current_report, current_date, prior_date)
    
    print(f"Report generated: {output_path}")
    
    # Distribution
    if src.config.EMAIL_TO:
        subject = "Financial Difficulty Modifications Report - Resolution Committee"
        body = """Hi,

Attached is the Financial Difficulty Modifications Report for the Resolution Committee Package. If you have any questions, please reach out to BusinessIntelligence@bcsbmail.com

Thanks!"""
        
        import cdutils.distribution # type: ignore
        cdutils.distribution.email_out(
            recipients=src.config.EMAIL_TO, 
            bcc_recipients=src.config.EMAIL_CC, 
            subject=subject, 
            body=body, 
            attachment_paths=[output_path]
        )
        print(f"Email sent to {len(src.config.EMAIL_TO)} recipients")
    else:
        print("Development mode - email not sent")


#################################



#################################
def report_creation(current_report, acctcommon, prior_report, prior_date_df, current_date, prior_date):
    """
    All core ETL & report creation is done here
    """
    acctcommon = acctcommon[['acctnbr','curracctstatcd']]
    
    current_report['Next Payment Due Date'] = pd.to_datetime(current_report['currduedate'], errors='coerce')
    current_report['TDR Date'] = pd.to_datetime(current_report['TDAT'], errors = 'coerce')
    
    current_report['TagType'] = 'Data'
    current_report['TagType'] = np.where(current_report['Product Name'].str.contains('Subtotal'),'Subtotal',current_report['TagType'])
    current_report['TagType'] = np.where((current_report['Product Name'].isna()) & (current_report['Net Balance'].notna()), 'Total Double', current_report['TagType'])

    current_report_cleaned = current_report.dropna(subset=['Account Number'])
    current_report_cleaned = current_report_cleaned[current_report_cleaned['Account Number'] != '']
    current_report_cleaned = current_report_cleaned.reset_index(drop=True)
    current_report_cleaned['Account Number'] = pd.to_numeric(current_report_cleaned['Account Number'], errors='coerce').astype('Int64')
    
    prior_report = prior_report.dropna(subset=['Account Number'])
    prior_report = prior_report[prior_report['Account Number'] != '']
    prior_report = prior_report.reset_index(drop=True)
    prior_report['Account Number'] = pd.to_numeric(prior_report['Account Number'], errors='coerce').astype('Int64')
    
    merged_df = pd.merge(current_report_cleaned, prior_report, how='outer', on='Account Number', suffixes=('_current','_prior'), indicator=True)
    
    new_items = merged_df[merged_df['_merge'] == 'left_only'].reset_index(drop=True)
    removed_items = merged_df[merged_df['_merge'] == 'right_only'].reset_index(drop=True)
    common_items = merged_df[merged_df['_merge'] == 'both'].reset_index(drop=True)

    numeric_columns = ['Account Number','Note Balance_current','Charged Off_current','Net Balance_current','Days Past Due_current','Note Balance_prior','Charged Off_prior','Net Balance_prior','Days Past Due_prior']
    merged_df[numeric_columns] = merged_df[numeric_columns].apply(pd.to_numeric, errors='coerce')
    
    current_total = merged_df['Net Balance_current'].sum()
    prior_total = merged_df['Net Balance_prior'].sum()
    
    try:
        new_items_grouped = new_items.groupby('Customer Name_current')['Net Balance_current'].sum()
        new_items_grouped = pd.DataFrame(new_items_grouped).reset_index()
        new_items_grouped.columns = ['Customer Name','Net Balance']
        new_items_grouped['TagType'] = "Data"
    except:
        new_items_grouped = None
        
    try:
        removed_merge = pd.merge(removed_items, acctcommon, left_on='Account Number', right_on='acctnbr', how='left', suffixes=('_removed','_acctcommon'))
        removed_merge = removed_merge.drop(columns='_merge')

        charged_off = removed_merge[removed_merge['curracctstatcd'] == 'CO']
        closed = removed_merge[removed_merge['curracctstatcd'] == 'CLS']
        removed_other = removed_merge[~removed_merge['curracctstatcd'].isin(['CO','CLS'])]

        removed_merge = pd.merge(removed_items, acctcommon, left_on='Account Number', right_on='acctnbr', how='left', suffixes=('_removed','_acctcommon'))
        removed_merge = removed_merge.drop(columns='_merge')
        charged_off = removed_merge[removed_merge['curracctstatcd'] == 'CO']
        closed = removed_merge[removed_merge['curracctstatcd'] == 'CLS']
        removed_other = removed_merge[~removed_merge['curracctstatcd'].isin(['CO','CLS'])]

        try:
            closed_grouped = closed.groupby('Customer Name_prior')['Net Balance_prior'].sum()*-1
            closed_grouped = pd.DataFrame(closed_grouped).reset_index()
            closed_grouped.columns = ['Customer Name','Net Balance']
            closed_grouped['TagType'] = "Data"
        except:
            closed_grouped = None

        try:
            charged_off_grouped = charged_off.groupby('Customer Name_prior')['Net Balance_prior'].sum()*-1
            charged_off_grouped = pd.DataFrame(charged_off_grouped).reset_index()
            charged_off_grouped.columns = ['Customer Name','Net Balance']
            charged_off_grouped['TagType'] = "Data"
        except:
            charged_off_grouped = None

        try:
            removed_other_grouped = removed_other.groupby('Customer Name_prior')['Net Balance_prior'].sum()*-1
            removed_other_grouped = pd.DataFrame(removed_other_grouped).reset_index()
            removed_other_grouped.columns = ['Customer Name','Net Balance']
            removed_other_grouped['TagType'] = "Data"
        except:
            removed_other_grouped = None
    except:
        closed_grouped = None
        charged_off_grouped = None
        removed_other_grouped = None
        
    common_items['Net Balance Diff'] = common_items['Net Balance_current'] - common_items['Net Balance_prior']
    common_items[['Net Balance_current','Net Balance_prior','Net Balance Diff']]

    loc_paydown = common_items[common_items['Net Balance Diff'] <= -25000]
    loc_advance = common_items[common_items['Net Balance Diff'] >= 25000]

    common_items['Charge Off Diff'] = common_items['Charged Off_current'] - common_items['Charged Off_prior']
    partial_charge_off = common_items[common_items['Charge Off Diff'] > 0]

    try:
        loc_paydown_grouped = loc_paydown.groupby('Customer Name_current')['Net Balance Diff'].sum()
        loc_paydown_grouped = pd.DataFrame(loc_paydown_grouped).reset_index()
        loc_paydown_grouped.columns = ['Customer Name','Net Balance']
        loc_paydown_grouped['Net Balance'] = pd.to_numeric(loc_paydown_grouped['Net Balance'], errors='coerce')
        loc_paydown_grouped['TagType'] = "Data"
    except:
        loc_paydown_grouped = None

    try:
        loc_advance_grouped = loc_advance.groupby('Customer Name_current')['Net Balance Diff'].sum()
        loc_advance_grouped = pd.DataFrame(loc_advance_grouped).reset_index()
        loc_advance_grouped.columns = ['Customer Name','Net Balance']
        loc_advance_grouped['TagType'] = "Data"
    except:
        loc_advance_grouped = None

    try:
        partial_charge_off_grouped = partial_charge_off.groupby('Customer Name_current')['Charge Off Diff'].sum()*-1
        partial_charge_off_grouped = pd.DataFrame(partial_charge_off_grouped).reset_index()
        partial_charge_off_grouped.columns = ['Customer Name','Net Balance']
        partial_charge_off_grouped['TagType'] = "Data"
    except:
        partial_charge_off_grouped = None
        
    def recon_helper(df, title_string):
        # Create title row with proper columns to match data structure
        title = {
            'Field1': [title_string], 
            'Customer Name': [''],
            'Net Balance': [np.nan],
            'TagType': ['Header']
        }
        title_df = pd.DataFrame(title)
        
        # Handle None DataFrames
        if df is None:
            # Return just the title row
            return title_df
        else:
            # Ensure df has Field1 column for consistency
            df_copy = df.copy()
            df_copy['Field1'] = df_copy.get('Customer Name', '')
            new_df = pd.concat([title_df, df_copy], ignore_index=True)
            return new_df
    
    new_items_grouped = recon_helper(new_items_grouped, "NEW ADDITIONS")
    closed_grouped = recon_helper(closed_grouped, "CLOSED")
    charged_off_grouped = recon_helper(charged_off_grouped, "CHARGED OFF")
    removed_other_grouped = recon_helper(removed_other_grouped, "REMOVED")
    loc_paydown_grouped = recon_helper(loc_paydown_grouped, "LOC PAYDOWN")
    loc_advance_grouped = recon_helper(loc_advance_grouped, "LOC ADVANCE")
    partial_charge_off_grouped = recon_helper(partial_charge_off_grouped, "PARTIAL CHARGE OFF")
    
    # Debug output for reconciliation components
    print("\n" + "="*80)
    print("RECONCILIATION DEBUGGING OUTPUT")
    print("="*80)
    
    # Calculate individual component sums
    new_items_sum = 0
    closed_sum = 0
    charged_off_sum = 0
    removed_other_sum = 0
    loc_paydown_sum = 0
    loc_advance_sum = 0
    partial_charge_off_sum = 0
    
    if new_items_grouped is not None and len(new_items_grouped) > 1:
        new_items_sum = new_items_grouped[new_items_grouped['TagType'] == 'Data']['Net Balance'].sum()
        print(f"NEW ADDITIONS: ${new_items_sum:,.2f}")
    
    if closed_grouped is not None and len(closed_grouped) > 1:
        closed_sum = closed_grouped[closed_grouped['TagType'] == 'Data']['Net Balance'].sum()
        print(f"CLOSED: ${closed_sum:,.2f}")
    
    if charged_off_grouped is not None and len(charged_off_grouped) > 1:
        charged_off_sum = charged_off_grouped[charged_off_grouped['TagType'] == 'Data']['Net Balance'].sum()
        print(f"CHARGED OFF: ${charged_off_sum:,.2f}")
    
    if removed_other_grouped is not None and len(removed_other_grouped) > 1:
        removed_other_sum = removed_other_grouped[removed_other_grouped['TagType'] == 'Data']['Net Balance'].sum()
        print(f"REMOVED: ${removed_other_sum:,.2f}")
    
    if loc_paydown_grouped is not None and len(loc_paydown_grouped) > 1:
        loc_paydown_sum = loc_paydown_grouped[loc_paydown_grouped['TagType'] == 'Data']['Net Balance'].sum()
        print(f"LOC PAYDOWN: ${loc_paydown_sum:,.2f}")
    
    if loc_advance_grouped is not None and len(loc_advance_grouped) > 1:
        loc_advance_sum = loc_advance_grouped[loc_advance_grouped['TagType'] == 'Data']['Net Balance'].sum()
        print(f"LOC ADVANCE: ${loc_advance_sum:,.2f}")
    
    if partial_charge_off_grouped is not None and len(partial_charge_off_grouped) > 1:
        partial_charge_off_sum = partial_charge_off_grouped[partial_charge_off_grouped['TagType'] == 'Data']['Net Balance'].sum()
        print(f"PARTIAL CHARGE OFF: ${partial_charge_off_sum:,.2f}")
    
    print("-"*80)
    
    try:
        reconciliation_body = pd.concat([new_items_grouped, closed_grouped, charged_off_grouped, removed_other_grouped, loc_paydown_grouped, loc_advance_grouped, partial_charge_off_grouped], ignore_index=True)
        # Filter out header rows when calculating sum
        data_rows = reconciliation_body[reconciliation_body['TagType'] != 'Header']
        recon_body_sum = float(data_rows['Net Balance'].sum())
    except Exception as e:
        print(f"Error calculating reconciliation body sum: {e}")
        recon_body_sum = 0

    total_net_adjustments = current_total - prior_total
    # Total Principal Payments = Net Adjustments - Sum of reconciliation components
    # This represents the unexplained portion of the change (e.g., regular payments, interest accruals, etc.)
    total_principal_payments = total_net_adjustments - recon_body_sum
    
    print(f"\nPrior Month Total: ${prior_total:,.2f}")
    print(f"Current Month Total: ${current_total:,.2f}")
    print(f"Total Net Adjustments (Current - Prior): ${total_net_adjustments:,.2f}")
    print(f"Sum of Reconciliation Components: ${recon_body_sum:,.2f}")
    print(f"Total Principal Payments (Net Adjustments - Components): ${total_principal_payments:,.2f}")
    
    # Verification calculation
    print(f"\nVERIFICATION:")
    print(f"Total Principal Payments = Net Adjustments - Sum of Components")
    print(f"                        = {total_net_adjustments:,.2f} - {recon_body_sum:,.2f}")
    print(f"                        = {total_principal_payments:,.2f}")
    print(f"\nThis represents the unexplained portion of the balance change")
    print(f"(e.g., regular principal payments, interest accruals, fees, etc.)")
    print("="*80 + "\n")
    
    
    reconciliation_title = {
        'Field1':["RECONCILIATION"],
        'Customer Name':[""],
        'Net Balance':[""],
        'TagType':["Header"]
    }
    reconciliation_title = pd.DataFrame(reconciliation_title)

    prior_month_title = {
        'Field1':[prior_date],
        'Customer Name':[""],
        'Net Balance':[prior_total],
        'TagType':["Total"]
    }
    prior_month_title = pd.DataFrame(prior_month_title)

    total_principal_payments_df = {
        'Field1':["TOTAL PRINCIPAL PAYMENTS"],
        'Customer Name':[""],
        'Net Balance':[total_principal_payments],
        'TagType':["Total"]
    }
    total_principal_payments_df = pd.DataFrame(total_principal_payments_df)

    total_net_adjustments_df = {
        'Field1':["TOTAL NET ADJUSTMENTS"],
        'Customer Name':[""],
        'Net Balance':[total_net_adjustments],
        'TagType':["Total"]
    }
    total_net_adjustments_df = pd.DataFrame(total_net_adjustments_df)

    current_month_title = {
        'Field1':[current_date],
        'Customer Name':[""],
        'Net Balance':[current_total],
        'TagType':["Total Double"]
    }
    current_month_title = pd.DataFrame(current_month_title)

    # Reconciliation Full
    dfs = [reconciliation_title, prior_month_title, reconciliation_body, total_principal_payments_df, total_net_adjustments_df, current_month_title]
    reconciliation_full = pd.concat(dfs, ignore_index=True)

    def dpd_section(df):
        bins = [0, 30, 60, float('inf')]
        labels = ['0-29','30-59','60+']
        
        df['bucket'] = pd.cut(df['Days Past Due'], bins=bins, labels=labels, right=False)
        
        summary = df.groupby(['bucket','Major'])['Account Number'].nunique().reset_index()
        summary.columns = ['Bucket','Major','Count']
        
        all_majors = ['CML','CNS','MTG']
        all_buckets = labels

        all_combinations = pd.MultiIndex.from_product([all_buckets, all_majors], names=['Bucket','Major']).to_frame(index=False)
        complete_summary = pd.merge(all_combinations, summary, on=['Bucket','Major'],how='left').fillna(0)

        complete_summary['Count'] = complete_summary['Count'].astype(int)
        
        cml_summary = complete_summary[complete_summary['Major'] == "CML"]
        cns_summary = complete_summary[complete_summary['Major'] == "CNS"]
        mtg_summary = complete_summary[complete_summary['Major'] == "MTG"]
        
        return cml_summary, cns_summary, mtg_summary


    def sum_section(df):
        bins = [0, 30, 60, float('inf')]
        labels = ['0-29','30-59','60+']
        
        df['bucket'] = pd.cut(df['Days Past Due'], bins=bins, labels=labels, right=False)
        
        summary = df.groupby(['bucket','Major'])['Net Balance'].sum().reset_index()
        summary.columns = ['Bucket','Major','Sum']
        
        all_majors = ['CML','CNS','MTG']
        all_buckets = labels

        all_combinations = pd.MultiIndex.from_product([all_buckets, all_majors], names=['Bucket','Major']).to_frame(index=False)
        complete_summary = pd.merge(all_combinations, summary, on=['Bucket','Major'],how='left').fillna(0)

        complete_summary['Sum'] = complete_summary['Sum'].astype(float)
        
        cml_summary = complete_summary[complete_summary['Major'] == "CML"]
        cns_summary = complete_summary[complete_summary['Major'] == "CNS"]
        mtg_summary = complete_summary[complete_summary['Major'] == "MTG"]
        
        return cml_summary, cns_summary, mtg_summary


    current_cml_sum, current_cns_sum, current_mtg_sum = sum_section(current_report_cleaned)
    
    prior_cml_sum, prior_cns_sum, prior_mtg_sum = sum_section(prior_report)

    cml_sum = pd.merge(current_cml_sum, prior_cml_sum, on='Bucket', how='inner', suffixes=('_current','_prior'))
    cns_sum = pd.merge(current_cns_sum, prior_cns_sum, on='Bucket', how='inner', suffixes=('_current','_prior'))
    mtg_sum = pd.merge(current_mtg_sum, prior_mtg_sum, on='Bucket', how='inner', suffixes=('_current','_prior'))
    
    current_cml_dpd, current_cns_dpd, current_mtg_dpd = dpd_section(current_report_cleaned)
    
    prior_cml_dpd, prior_cns_dpd, prior_mtg_dpd = dpd_section(prior_report)

    cml_sum = pd.merge(cml_sum, current_cml_dpd, on='Bucket', how='inner')
    cns_sum = pd.merge(cns_sum, current_cns_dpd, on='Bucket', how='inner')
    mtg_sum = pd.merge(mtg_sum, current_mtg_dpd, on='Bucket', how='inner')
    
    cml_sum = cml_sum[['Bucket','Sum_current','Sum_prior','Count']]
    cns_sum = cns_sum[['Bucket','Sum_current','Sum_prior','Count']]
    mtg_sum = mtg_sum[['Bucket','Sum_current','Sum_prior','Count']]

    def summary_section_completed(df, major_type):
        sum_current = df['Sum_current'].sum()
        sum_prior = df['Sum_prior'].sum()
        count = df['Count'].sum()

        data = {
            'Bucket':[""],
            'Sum_current':[sum_current],
            'Sum_prior':[sum_prior],
            'Count':[count]
        }
        data = pd.DataFrame(data)
        new_df = pd.concat([df, data])
        new_df['Field2'] = None
        new_df = new_df[['Bucket','Field2','Sum_current','Sum_prior','Count']]

        # Renaming Rows & Adding TagType based on major type
        if major_type == 'CML':
            new_df['Bucket'] = np.where(new_df['Bucket'] == "0-29","TOTAL COMMERCIAL 0-29 DPD", new_df['Bucket'])
            new_df['Bucket'] = np.where(new_df['Bucket'] == "30-59","TOTAL COMMERCIAL 30-59 DPD", new_df['Bucket'])
            new_df['Bucket'] = np.where(new_df['Bucket'] == "60+","TOTAL COMMERCIAL 60+ DPD", new_df['Bucket'])
        elif major_type == 'CNS':
            new_df['Bucket'] = np.where(new_df['Bucket'] == "0-29","TOTAL CONSUMER 0-29 DPD", new_df['Bucket'])
            new_df['Bucket'] = np.where(new_df['Bucket'] == "30-59","TOTAL CONSUMER 30-59 DPD", new_df['Bucket'])
            new_df['Bucket'] = np.where(new_df['Bucket'] == "60+","TOTAL CONSUMER 60+ DPD", new_df['Bucket'])
        elif major_type == 'MTG':
            new_df['Bucket'] = np.where(new_df['Bucket'] == "0-29","TOTAL RESIDENTIAL 0-29 DPD", new_df['Bucket'])
            new_df['Bucket'] = np.where(new_df['Bucket'] == "30-59","TOTAL RESIDENTIAL 30-59 DPD", new_df['Bucket'])
            new_df['Bucket'] = np.where(new_df['Bucket'] == "60+","TOTAL RESIDENTIAL 60+ DPD", new_df['Bucket'])
        
        new_df['TagType'] = np.where(new_df['Bucket'] == "", "Total Double", "Data")

        return new_df


    completed_summary_cml = summary_section_completed(cml_sum, 'CML')
    completed_summary_cns = summary_section_completed(cns_sum, 'CNS')
    completed_summary_mtg = summary_section_completed(mtg_sum, 'MTG')

    empty_row = pd.DataFrame([None] * len(completed_summary_cml.columns)).T
    empty_row.columns = completed_summary_cml.columns
    summary_combined_df = pd.concat([completed_summary_cml, empty_row, completed_summary_cns, empty_row, completed_summary_mtg], ignore_index=True)

    # Adding title and dates to Summary section
    summary_title_row = {
        'Bucket':['FDM LOAN DELINQUENCY SUMMARY'],
        'Field2':None,
        'Sum_current':None,
        'Sum_prior':None,
        'Count':None,
        'TagType':'Header'
    }
    summary_title_row_df = pd.DataFrame(summary_title_row)

    date_row_summary = {
        'Bucket':None,
        'Field2':None,
        'Sum_current':[current_date],
        'Sum_prior':[prior_date],
        'Count':None,
        'TagType':'Total Date'
    }

    date_row_summary_df = pd.DataFrame(date_row_summary)


    summary_combined_df = pd.concat([summary_title_row_df, date_row_summary_df, summary_combined_df], ignore_index=True)
    grand_sum_summary_df = summary_combined_df[summary_combined_df['TagType'] == 'Total Double']
    sum_current_all = grand_sum_summary_df['Sum_current'].sum()
    sum_prior_all = grand_sum_summary_df['Sum_prior'].sum()
    count_current_all = grand_sum_summary_df['Count'].sum()

    grand_sum_data = {
        'Bucket':None,
        'Field2':None,
        'Sum_current':[sum_current_all],
        'Sum_prior':[sum_prior_all],
        'Count':[count_current_all],
        'TagType':'Total Double'
    }

    grand_sum_data_df = pd.DataFrame(grand_sum_data)

    summary_combined_df = pd.concat([summary_combined_df, grand_sum_data_df], ignore_index=True)
    
    return reconciliation_full, summary_combined_df, current_report, current_date, prior_date


def output_to_excel(output_dir, filename, reconciliation_full, summary_combined_df, current_report, current_date, prior_date):
    """
    Output to Excel with formatting and styles applied
    """
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    output_file = output_path / filename
    
    if 'Net Balance' in reconciliation_full.columns:
        reconciliation_full['Net Balance'] = pd.to_numeric(reconciliation_full['Net Balance'], errors='coerce')

    with pd.ExcelWriter(str(output_file), engine='openpyxl', mode='w') as writer:

        reconciliation_full.to_excel(writer, index=False, sheet_name='Sheet1', startcol=4, startrow=5, header=False)

        start_row = 5 + len(reconciliation_full) + 1
        summary_section_start = start_row + 1

        summary_combined_df.to_excel(writer, index=False, sheet_name='Sheet1', startcol=4, startrow=start_row, header=False)

        start_row = start_row + len(summary_combined_df) + 1
        report_header_row = start_row + 1

        # Clean up the report columns and add subtotals
        report_columns = [
            'Product Name', 'Risk', 'Customer Name', 'Account Number', 
            'Net Balance', 'Days Past Due', 'Next Payment Due Date', 
            'Non Accrual', 'TDR Date', 'TDR Note'
        ]
        
        # Create a copy of current_report and ensure all required columns exist
        current_report_clean = current_report.copy()
        
        # Add missing columns if they don't exist
        if 'Risk' not in current_report_clean.columns:
            current_report_clean['Risk'] = ''
        if 'Non Accrual' not in current_report_clean.columns:
            current_report_clean['Non Accrual'] = ''
        if 'TDR Note' not in current_report_clean.columns:
            current_report_clean['TDR Note'] = current_report_clean.get('TNOT', '')
        
        # Map existing columns to the desired names
        column_mapping = {
            'risk': 'Risk',
            'TNOT': 'TDR Note'
        }
        
        # Rename columns if they exist with different names
        for old_name, new_name in column_mapping.items():
            if old_name in current_report_clean.columns and new_name not in current_report_clean.columns:
                current_report_clean[new_name] = current_report_clean[old_name]
        
        # Format the data properly
        # Risk rating - ensure it's populated
        if 'risk' in current_report_clean.columns:
            current_report_clean['Risk'] = current_report_clean['risk'].fillna('')
        
        # Days Past Due - ensure it's a number, not currency
        current_report_clean['Days Past Due'] = pd.to_numeric(current_report_clean['Days Past Due'], errors='coerce').fillna(0)
        
        # Next Payment Due Date - format as date only
        current_report_clean['Next Payment Due Date'] = pd.to_datetime(current_report_clean['Next Payment Due Date'], errors='coerce').dt.date
        
        # Non Accrual - convert to Yes/No
        current_report_clean['Non Accrual'] = current_report_clean['Non Accrual'].fillna('No')
        current_report_clean['Non Accrual'] = current_report_clean['Non Accrual'].map({'Y': 'Yes', 'N': 'No', '': 'No'}).fillna('No')
        
        # TDR Date - format as date only
        current_report_clean['TDR Date'] = pd.to_datetime(current_report_clean['TDR Date'], errors='coerce').dt.date
        
        # Select and reorder columns for the report
        current_report_clean = current_report_clean[report_columns].copy()
        
        # Create structured report with subtotals and grouping
        final_report_rows = []
        group_labels = [('CML', 'Commercial'), ('CNS', 'Consumer'), ('MTG', 'Residential')]
        total_count = 0
        total_amount = Decimal('0.0')
        for code, label in group_labels:
            group_data = current_report_clean[current_report['Major'] == code]
            if not group_data.empty:
                # Sort by Risk rating (4's first, then 5's, then others)
                group_data = group_data.sort_values('Risk', key=lambda x: x.astype(str))
                final_report_rows.append(group_data)
                group_count = len(group_data)
                # Ensure group_sum is Decimal
                group_sum = sum([Decimal(str(x)) for x in group_data['Net Balance']])
                total_count += group_count
                total_amount += group_sum
                subtotal_row = pd.DataFrame([{
                    'Product Name': f'Subtotal: {label}',
                    'Risk': '',
                    'Customer Name': '',
                    'Account Number': group_count,
                    'Net Balance': float(group_sum),
                    'Days Past Due': '',
                    'Next Payment Due Date': '',
                    'Non Accrual': '',
                    'TDR Date': '',
                    'TDR Note': ''
                }])
                final_report_rows.append(subtotal_row)
        # Grand total row
        grand_total = pd.DataFrame([{
            'Product Name': 'Total',
            'Risk': '',
            'Customer Name': '',
            'Account Number': total_count,
            'Net Balance': float(total_amount),
            'Days Past Due': '',
            'Next Payment Due Date': '',
            'Non Accrual': '',
            'TDR Date': '',
            'TDR Note': ''
        }])
        final_report_rows.append(grand_total)
        # Combine all rows
        final_report = pd.concat(final_report_rows, ignore_index=True)
        final_report.to_excel(writer, index=False, sheet_name='Sheet1', startcol=3, startrow=start_row, header=True)

    wb = load_workbook(str(output_file))
    ws = wb['Sheet1']

    # Dimensions - Set more appropriate column widths
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['B'].width = 1
    ws.column_dimensions['C'].width = 1
    ws.column_dimensions['D'].width = 27  # Reconciliation section text
    ws.column_dimensions['E'].width = 6   # Risk column (small, will allow overflow)
    ws.column_dimensions['F'].width = 35  # Customer Name
    ws.column_dimensions['G'].width = 14  # Account Number
    ws.column_dimensions['H'].width = 15  # Net Balance
    ws.column_dimensions['I'].width = 10  # Days Past Due
    ws.column_dimensions['J'].width = 15  # Next Payment Due Date
    ws.column_dimensions['K'].width = 12  # Non Accrual
    ws.column_dimensions['L'].width = 12  # TDR Date
    ws.column_dimensions['M'].width = 50  # TDR Note (increased width for multi-line text)
    ws.column_dimensions['N'].width = 1   # Empty column
    ws.column_dimensions['O'].width = 1   # Empty column

    # Cell Font & Formatting
    upper_section_font = Font(size=10, name='Arial', bold=True)
    bold_italics_font = Font(size=10, name='Arial', bold=True, italic=True)
    subtitle_font = Font(size=10, name='Arial', bold=True)
    data_font = Font(size=10, name='Arial')
    header_font = Font(size=9, name='Arial', bold=True)  # Smaller font for report headers
    wrap_alignment = Alignment(wrap_text=True, vertical='bottom', horizontal='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    title_block = Font(size=14, name='Arial', bold=True)
    accounting_style = '"$"#,##0.00_);("$"#,##0.00)'
    count_format = '#,##0_);(#,##0)'
    parentheses_currency = '"$"#,##0.00_);("$"#,##0.00)'

    # Borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thin'))
    double_bottom_border= Border(bottom=Side(style='double'), top=Side(style='thin'))
    side_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    top_border = Border(top=Side(style='thin'))
    top_and_bottom_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))


    # Header
    ws['E1'] = "BRISTOL COUNTY SAVINGS BANK"
    ws['E1'].font = upper_section_font
    ws['E2'] = "Financial Difficulty Modifications (FDM) Report"
    ws['E2'].font = upper_section_font
    ws['E3'] = f"{current_date}"
    ws['E3'].font = upper_section_font

    # Saving Last row as a variable
    last_row = ws.max_row

    # Bolding cells
    for row in range(1, report_header_row):
        cell = ws.cell(row=row, column=5)
        cell.font = upper_section_font

    # Wrap Text and format header row of report
    for row in ws.iter_rows(min_row=report_header_row, max_row=report_header_row):
        for cell in row:
            if cell.column >= 4:  # Only format report columns
                cell.alignment = wrap_alignment
                cell.font = header_font  # Use smaller header font

    # Reconciliation Formatting
    # Ensure all Net Balance cells in reconciliation section use parentheses for negatives
    for row in range(1, summary_section_start-1):
        for col in ['F', 'G']:
            cell = ws[f'{col}{row}']
            try:
                val = float(cell.value)
                cell.number_format = parentheses_currency if val < 0 else accounting_style
            except (TypeError, ValueError):
                pass

    for row in range(1, summary_section_start-1):
        for col in ['H']:
            cell = ws[f'{col}{row}']
            cell.value = None

    # Summary Formatting
    for row in ws.iter_rows(min_col=10, max_col=10, min_row=summary_section_start, max_row=report_header_row-1):
        for cell in row:
            row_idx = cell.row

            if cell.value == 'Header':
                for col in ws.iter_cols(min_row=row_idx, max_row=row_idx):
                    for cell2 in col:
                        cell2.font = upper_section_font

            elif cell.value == 'Total Date':
                for col in ws.iter_cols(min_row=row_idx, max_row=row_idx, min_col=7, max_col=9):
                    for cell2 in col:
                        cell2.font = upper_section_font
                        cell2.border = bottom_border
                        cell2.alignment = center_alignment

            elif cell.value == 'Total Double':
                for col in ws.iter_cols(min_row=row_idx, max_row=row_idx, min_col=7, max_col=9):
                    for cell2 in col:
                        cell2.border = double_bottom_border

    for row in range(summary_section_start+2, report_header_row-1):
        for col in ['G','H']:
            cell = ws[f'{col}{row}']
            try:
                val = float(cell.value)
                cell.number_format = parentheses_currency if val < 0 else accounting_style
            except (TypeError, ValueError):
                pass

        for col in ['I']:
            cell = ws[f'{col}{row}']
            cell.alignment = center_alignment
            cell.number_format = count_format  # Format count column properly

    for row in range(summary_section_start, report_header_row-1):
        for col in ['J']:
            cell = ws[f'{col}{row}']
            cell.value = None

    # Report Formatting - Apply bold italics to entire subtotal and total rows
    for row_num in range(report_header_row + 1, last_row + 1):
        try:
            # Check if this is a subtotal or total row by looking at Product Name column (column D = 4)
            product_name_value = ws.cell(row=row_num, column=4).value
            if product_name_value and ('Subtotal:' in str(product_name_value) or product_name_value == 'Total'):
                # Apply bold italics font to all cells in this row
                for col in range(4, 14):  # Columns D through M
                    ws.cell(row=row_num, column=col).font = bold_italics_font
        except Exception:
            pass

    # Report data formatting with improved alignment
    for row in ws.iter_rows(min_row=report_header_row + 1, max_row=last_row, min_col=4, max_col=13):
        for cell in row:
            col_letter = cell.column_letter
            
            # Set alignment based on column type
            if col_letter in ['D']:  # Product Name - left align
                cell.alignment = left_alignment
            elif col_letter in ['E']:  # Risk - center align
                cell.alignment = center_alignment
            elif col_letter in ['F']:  # Customer Name - left align
                cell.alignment = left_alignment
            elif col_letter in ['G', 'H', 'I', 'J', 'K']:  # Numbers and dates - right align
                cell.alignment = right_alignment
            elif col_letter == 'L':  # TDR Date - left align
                cell.alignment = left_alignment
            elif col_letter == 'M':  # TDR Note - left align with text wrapping
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Apply data font to all cells
            if not cell.font.bold:  # Don't override bold formatting for subtotals
                cell.font = data_font

    # Format Account Number column (column G) 
    for row in ws.iter_rows(min_row=report_header_row + 1, max_row=last_row, min_col=7, max_col=7):
        for cell in row:
            try:
                val = float(cell.value)
                cell.number_format = '0'  # No commas or decimal places for account numbers
                cell.alignment = right_alignment
            except (TypeError, ValueError):
                pass
    
    # Format Net Balance as currency (column H)
    for row in ws.iter_rows(min_row=report_header_row + 1, max_row=last_row, min_col=8, max_col=8):
        for cell in row:
            try:
                val = float(cell.value)
                cell.number_format = parentheses_currency if val < 0 else accounting_style
                cell.alignment = right_alignment
            except (TypeError, ValueError):
                pass
    
    # Format Days Past Due as number (column I)
    for row in ws.iter_rows(min_row=report_header_row + 1, max_row=last_row, min_col=9, max_col=9):
        for cell in row:
            try:
                val = float(cell.value)
                cell.number_format = '#,##0'
                cell.alignment = right_alignment
            except (TypeError, ValueError):
                pass
    
    # Format Net Balance as currency (column J - this was moved from column 10 due to column shifts)
    for row in ws.iter_rows(min_row=report_header_row + 1, max_row=last_row, min_col=10, max_col=10):
        for cell in row:
            try:
                val = float(cell.value)
                cell.number_format = parentheses_currency if val < 0 else accounting_style
                cell.alignment = right_alignment
            except (TypeError, ValueError):
                pass

    # Smart auto-width for specific columns only
    # Only auto-size columns that benefit from dynamic sizing
    auto_size_columns = ['D', 'F', 'L', 'M']  # Reconciliation text, Customer Name, TDR Date, TDR Note
    
    for col_letter in auto_size_columns:
        max_length = 0
        col = ws[col_letter]
        
        for cell in col:
            try:
                if cell.value and str(cell.value).strip():
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        
        # Set minimum and maximum widths based on column type
        if col_letter == 'D':  # Reconciliation section
            ws.column_dimensions[col_letter].width = max(25, min(max_length + 2, 50))
        elif col_letter == 'F':  # Customer Name
            ws.column_dimensions[col_letter].width = max(20, min(max_length + 2, 40))
        elif col_letter == 'L':  # TDR Date
            ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 25))
        elif col_letter == 'M':  # TDR Note - keep fixed width for wrapping
            ws.column_dimensions[col_letter].width = 50

    # Remove the old auto-width section and column R cleanup
    for row in range(report_header_row, last_row+1):
        for col in ['R']:
            cell = ws[f'{col}{row}']
            cell.value = None
            cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

    wb.save(str(output_file))


if __name__ == "__main__":
    main()
    print("Complete!")