from sqlalchemy import text
import sys
import os
import cdutils.database.connect # type: ignore 

import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment

import numpy as np
from pathlib import Path
import pandas as pd
from datetime import datetime


import src.config

"""
Classifieds Report - Main Entry Point

BUSINESS LOGIC EXTRACTED:

Data Sources & Relationships:
- COCCDM.WH_ACCTCOMMON: Account master data with status, balance, and product information
- COCCDM.WH_ACCTLOAN: Risk ratings, credit limits, charge-offs, and payment due dates
- COCCDM.WH_LOANS: Original loan terms and FDIC category descriptions
- COCCDM.WH_ACCT: Maturity dates and NAICS descriptions for commercial loans
- OSIBANK.WH_ORG/WH_PERS: NAICS industry codes for commercial borrowers
- OSIBANK.WH_PROP2: Property descriptions for mortgage loans

Business Rules:
- Focuses on loans with risk ratings 4 (Special Mention) and 5 (Substandard)
- Grade 4 Commercial includes both CML and MLN (Money Line) account types
- Net Balance = Book Balance - Charged Off Amount
- Net Available = Available Balance × (1 - Percent Sold)
- Total Exposure = Net Balance + Net Available + Net Collateral Reserve
- FDIC categories mapped to portfolio buckets (CRE, C&I, Consumer, etc.)

Data Processing Flow:
1. Fetch current and prior month-end snapshots for two-period comparison
2. Apply business rules for exposure calculations and portfolio categorization
3. Filter to classified loans (risk ratings 4 and 5) by major account type
4. Build reconciliation showing new loans, risk rating changes, payoffs, and advances
5. Generate summary sections with account details sorted by customer name and days past due
6. Format Excel output with reconciliation boxes and detailed loan listings

Key Calculations:
- Total Exposure = Net Balance + Net Available + Net Collateral Reserve
- Days Past Due = Report Date - Next Payment Due Date (minimum 0)
- Month-over-month balance reconciliation with detailed change tracking
- Net payments and advances as balancing reconciliation items

Business Intelligence Value:
Provides comprehensive classified loan analysis for regulatory compliance and Resolution
Committee oversight, tracking risk rating migrations and exposure changes with detailed
reconciliation capabilities.
"""

def fetch_data_for_date(month_end: str = "2025-01-31"):
    """
    Pull WH_ACCT* snapshots for a single month-end.

    Parameters
    ----------
    month_end : str
        The month-end you want, **YYYY-MM-DD** (no time).
        Example: "2025-01-31"
    """

    # Oracle wants a full timestamp; append "00:00:00"
    effdate_sql = f"TO_DATE('{month_end} 00:00:00', 'YYYY-MM-DD HH24:MI:SS')"

    wh_acctcommon = text(f"""
        SELECT
            a.EFFDATE,
            a.ACCTNBR,
            a.OWNERSORTNAME,
            a.PRODUCT,
            a.NOTEOPENAMT,
            a.RATETYPCD,
            a.MJACCTTYPCD,
            a.CURRMIACCTTYPCD,
            a.CURRACCTSTATCD,
            a.NOTEINTRATE,
            a.BOOKBALANCE,
            a.NOTEBAL,
            a.LOANOFFICER,
            a.ACCTOFFICER,
            a.TAXRPTFORORGNBR,
            a.TAXRPTFORPERSNBR
        FROM COCCDM.WH_ACCTCOMMON a
        WHERE a.CURRACCTSTATCD IN ('ACT','NPFM','CO','CLS')
        AND a.EFFDATE = {effdate_sql}
    """)

    wh_loans = text(f"""
        SELECT
            a.ACCTNBR,
            a.ORIGDATE,
            a.CURRTERM,
            a.LOANIDX,
            a.RCF,
            a.AVAILBALAMT,
            a.FDICCATDESC,
            a.ORIGBAL
        FROM COCCDM.WH_LOANS a
        WHERE a.RUNDATE = {effdate_sql}
    """)

    wh_acctloan = text(f"""
        SELECT
            a.ACCTNBR,
            a.CREDITLIMITAMT,
            a.ORIGINTRATE,
            a.MARGINFIXED,
            a.FDICCATCD,
            a.AMORTTERM,
            a.TOTALPCTSOLD,
            a.COBAL,
            a.CREDLIMITCLATRESAMT,
            a.RISKRATINGCD,
            a.NEXTDUEDATE,
            a.CURRDUEDATE
        FROM COCCDM.WH_ACCTLOAN a
        WHERE a.EFFDATE = {effdate_sql}
    """)

    wh_acct = text(f"""
        SELECT
            a.ACCTNBR,
            a.DATEMAT,
            a.NAICSDESC
        FROM COCCDM.WH_ACCT a
        WHERE a.RUNDATE = {effdate_sql}
    """)
    
    wh_org = text(f"""
    SELECT ORGNBR, NAICSCD, NAICSCDDESC AS ORGNAICS, ADDDATE
    FROM OSIBANK.WH_ORG
    """)

    wh_pers = text(f"""
    SELECT PERSNBR, NAICSDESC AS PERSNAICS, NAICSCD
    FROM OSIBANK.WH_PERS
    """)

    wh_prop2 = text(f"""
    SELECT PROPDESC, ACCTNBR, PROPNBR, EFFDATE, RUNDATE
    FROM OSIBANK.WH_PROP2

    """)

    queries = [
        {'key': 'wh_acctcommon', 'sql': wh_acctcommon, 'engine': 2},
        {'key': 'wh_loans',      'sql': wh_loans,      'engine': 2},
        {'key': 'wh_acctloan',   'sql': wh_acctloan,   'engine': 2},
        {'key': 'wh_acct',       'sql': wh_acct,       'engine': 2},
        {'key': 'org',     'sql': wh_org,     'engine': 1},
        {'key': 'pers',    'sql': wh_pers,    'engine': 1},
        {'key': 'prop2',   'sql': wh_prop2,   'engine': 1}
    ]

    return cdutils.database.connect.retrieve_data(queries)

def main():
    """Main report execution function"""
    print("Starting Resolution Committee Classifieds Report")
    


    def get_last_2_month_ends():
        ends = pd.date_range(end=pd.Timestamp.today(), periods=2, freq="BME")
        return [d.to_pydatetime().date() for d in ends]

    LAST_2_MONTH_ENDS = get_last_2_month_ends()
    prev_end = str(LAST_2_MONTH_ENDS[0])
    curr_end = str(LAST_2_MONTH_ENDS[1])

    curr_end_xlsx = curr_end + "classifieds_Report.xlsx"
    output_path = src.config.OUTPUT_DIR / curr_end_xlsx

    data_curr  = fetch_data_for_date(curr_end)

    data_prev  = fetch_data_for_date(prev_end)


    def _force_string(df: pd.DataFrame, col: str) -> pd.DataFrame:
        if col not in df.columns:
            df[col] = None            # create if missing
        df[col] = df[col].astype(str).str.strip()
        return df


    # ------------------------------------------------------------------
    def _transform_snapshot(data_dict: dict) -> pd.DataFrame:
        # ---------- unpack ----------
        wh_acctcommon = data_dict["wh_acctcommon"].copy()
        wh_loans      = data_dict["wh_loans"].copy()
        wh_acctloan   = data_dict["wh_acctloan"].copy()
        wh_acct       = data_dict["wh_acct"].copy()


        for df in (wh_acctcommon, wh_loans, wh_acctloan, wh_acct):
            _force_string(df, "acctnbr")

        # ---------- joins ----------
        loans = (
            wh_acctcommon
            .merge(wh_acctloan, on="acctnbr", how="left")
            .merge(wh_loans,      on="acctnbr", how="left")
            .merge(wh_acct,       on="acctnbr", how="left")
        )

        # ---------- numeric columns used in the maths ----------
        num_cols = [
            "bookbalance", "notebal",
            "availbalamt", "totalpctsold",
            "noteopenamt", "creditlimitamt",
            "noteintrate", "cobal", "credlimitclatresamt",
        ]
        for col in num_cols:
            if col not in loans.columns:
                loans[col] = 0.0
            loans[col] = pd.to_numeric(loans[col], errors="coerce")
            loans[col] = loans[col].fillna(0.0)

        # ---------- exposure mathematica ----------
        # Tax-exempt bonds (CM45) use NOTEBAL instead of BOOKBALANCE
        loans["bookbalance"] = np.where(
            loans["currmiaccttypcd"] == "CM45",
            loans["notebal"],
            loans["bookbalance"],
        )
        loans["Net Balance"]           = loans["bookbalance"] - loans["cobal"]
        loans["Net Available"]         = loans["availbalamt"] * (1 - loans["totalpctsold"])
        loans["Net Collateral Reserve"] = (
            loans["credlimitclatresamt"] * (1 - loans["totalpctsold"])
        )
        loans["Total Exposure"] = (
            loans["Net Balance"] + loans["Net Available"] + loans["Net Collateral Reserve"]
        )


        loans["origdate"] = pd.to_datetime(loans["origdate"], errors="coerce")
        loans["orig_ttl_loan_amt"] = np.where(
            loans["noteopenamt"] == 0, loans["creditlimitamt"], loans["noteopenamt"]
        )


        # step 1: redefine fdiccatcd for special cases
        loans.loc[loans["currmiaccttypcd"].isin(["CM15", "CM16"]), "fdiccatcd"] = "AUTO"
        loans.loc[loans["currmiaccttypcd"].isin(["CM46", "CM47"]), "fdiccatcd"] = "HOA"
        loans.loc[loans["currmiaccttypcd"] == "CM45",              "fdiccatcd"] = "OTAL"
        loans.loc[loans["mjaccttypcd"] == "MTG",                    "fdiccatcd"] = "MTG"
        loans.loc[loans["currmiaccttypcd"].isin(["IL09", "IL10"]), "fdiccatcd"] = "CNOT"
        loans["fdiccatcd"] = loans["fdiccatcd"].fillna("OTAL")

        # step 2: final portfolio buckets
        fdic_groups = {
            "CRE":        ["CNFM","OTCN","LAND","LNDV","RECN","REFI","REOE","REJU",
                        "REOW","RENO","REMU","OTAL","AGPR","REFM","LENO"],
            "C&I":        ["CIUS"],
            "HOA":        ["HOA"],
            "Residential":["MTG"],
            "Consumer":   ["CNOT","CNCR"],
            "Indirect":   ["AUTO"],
        }
        code_to_group = {code: grp for grp, codes in fdic_groups.items() for code in codes}
        loans["Category"] = loans["fdiccatcd"].map(code_to_group)


        loans = loans[loans["mjaccttypcd"].isin(["CML", "MLN", "MTG", "CNS"])]
        loans = loans.sort_values(["Total Exposure", "acctnbr"], ascending=[False, True])
        
        
        org_df = data_dict.get("org")
        pers_df = data_dict.get("pers")
        prop2_df = data_dict.get("prop2")

        if org_df is not None:
            org_df['adddate'] = pd.to_datetime(org_df['adddate'], errors="coerce")
            org_df = org_df.loc[org_df.groupby('orgnbr')['adddate'].idxmax()]  # clean_duplicates

        # Rename required columns in main loans df
        loans.rename(columns={
            "taxrptfororgnbr": "orgnbr",
            "taxrptforpersnbr": "persnbr"
        }, inplace=True)

        # Merge in org and pers
        if org_df is not None:
            loans = loans.merge(org_df[["orgnbr", "orgnaics"]], on="orgnbr", how="left")

        if pers_df is not None:
            loans = loans.merge(pers_df[["persnbr", "persnaics"]], on="persnbr", how="left")

        # Merge in prop2 for MTG loans only
        if prop2_df is not None:
            _force_string(prop2_df, "acctnbr")
            loans = loans.merge(prop2_df[["acctnbr", "propdesc"]], on="acctnbr", how="left")

        # Final field logic - prioritize wh_acct NAICSDESC for commercial loans
        # For commercial loans (non-MTG), prioritize NAICSDESC from wh_acct if not null
        commercial_mask = loans["mjaccttypcd"] != "MTG"
        loans.loc[commercial_mask, "naicsdesc"] = loans.loc[commercial_mask, "naicsdesc"].combine_first(
            loans.loc[commercial_mask, "orgnaics"].combine_first(loans.loc[commercial_mask, "persnaics"])
        )
        
        # For MTG loans, use propdesc as before
        loans.loc[loans["mjaccttypcd"] == "MTG", "naicsdesc"] = loans.loc[loans["mjaccttypcd"] == "MTG", "propdesc"]

        # Final formatting
        loans["officer"] = loans["loanofficer"].fillna(loans["acctofficer"])
        # Use currduedate if available, else fallback to nextduedate
        loans["nextpaymentduedate"] = loans["currduedate"].combine_first(loans["nextduedate"])

        
        def calculate_non_accrual_and_dpd(loans, curr_end_date):
            loans["curr_end_date"] = pd.to_datetime(curr_end_date)
            loans["nextpaymentduedate"] = pd.to_datetime(loans["nextpaymentduedate"], errors='coerce')

            loans["nonaccrual"] = loans["curracctstatcd"].apply(lambda x: 'Yes' if x == 'NPFM' else 'No')
            loans["dpd"] = (loans["curr_end_date"] - loans["nextpaymentduedate"]).dt.days
            loans.loc[loans["dpd"] < 0, "dpd"] = 0  
            
            # Edge case fix: If Net Balance is 0, set DPD to 0
            loans.loc[loans["Net Balance"] == 0, "dpd"] = 0
            
            return loans

        loans = calculate_non_accrual_and_dpd(loans, curr_end)

        return loans


    df_prev = _transform_snapshot(data_prev)   # <- prev_end (2025-04-30)
    df_curr = _transform_snapshot(data_curr)   # <- curr_end (2025-03-28)


    # --- Raw input copies

    df_prev_raw = df_prev
    df_curr_raw = df_curr

    # --- Collapse to latest effdate per acctnbr
    def collapse_to_latest(df):
        df = df.copy()
        df['effdate'] = pd.to_datetime(df['effdate'])
        df = df.sort_values(['acctnbr', 'effdate'], ascending=[True, False])
        return df.drop_duplicates('acctnbr', keep='first')

    df_prev = collapse_to_latest(df_prev_raw)
    df_curr = collapse_to_latest(df_curr_raw)

    # --- Categories
    cats = {
        "4-CML": ("CML", "4"),
        "5-CML": ("CML", "5"),
        "5-MTG": ("MTG", "5"),
    }

    # --- Column mapping
    column_mapping = {
        "product": "Product Name",
        "officer": "Responsibility Officer",
        "riskratingcd": "Risk Rating",
        "ownersortname": "Customer Name",
        "acctnbr": "Account\nNumber",
        "mjaccttypcd": "Major",
        "notebal": "Current\nBalance",
        "cobal": "Charge\nOff",
        "Net Balance": "Net\nBalance",
        #"notebal": "Current Note Balance",
        "nonaccrual": "Non\nAccrual",
        "dpd": "Days Past\nDue",
        "nextpaymentduedate": "Next\nPayment\nDue Date",
        "naicsdesc": "NAICS Description",
        "effdate": "EFFDATE",
        
    }

    def format_for_export(df):
        out = pd.DataFrame()
        for raw, nice in column_mapping.items():
            out[nice] = df[raw] if raw in df.columns else ""
        return out

    # --- Reconciliation builder
    def build_reconciliation(prev_df, curr_df, cat_label):
        p = prev_df.copy()
        c = curr_df.copy()
        full_curr = df_curr.copy()
        rows = []

        start_bal = p["Net Balance"].sum()
        end_bal = c["Net Balance"].sum()
        start_date = p["effdate"].iloc[0].strftime("%m/%d/%Y") if not p.empty else "N/A"
        end_date = c["effdate"].iloc[0].strftime("%m/%d/%Y") if not c.empty else "N/A"

        prev_rr = cat_label.split("-")[0]

        rows.append([f"CML Grade {prev_rr}" if "CML" in cat_label else cat_label, "Customer Name", "Current RR", "$", "#"])
        rows.append([start_date, "", "", round(start_bal, 2), len(p)])

        # Helper to group by customer
        def group_by_customer(df, name_col, amt_col):
            grouped = df.groupby(name_col).agg(
                {amt_col: "sum", "acctnbr": "count"}
            ).reset_index()
            return grouped

        # New Loans
        new_loans = c[~c["acctnbr"].isin(p["acctnbr"])]
        if not new_loans.empty:
            rows.append([f"New Grade {prev_rr}", "", "", "", ""])
            grouped = group_by_customer(new_loans, "ownersortname", "Net Balance")
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname"], "", r["Net Balance"], r["acctnbr"]])

        # Risk Rating Change
        p_matched = p[p["acctnbr"].isin(full_curr["acctnbr"])]
        joined = pd.merge(p_matched, full_curr, on="acctnbr", suffixes=("_p", "_c"))
        rr_change = joined[(joined["riskratingcd_p"] == prev_rr) & (joined["riskratingcd_c"] != prev_rr)]
        if not rr_change.empty:
            rows.append([f"Risk Rating Change (Previously Rated {prev_rr})", "", "", "", ""])
            grouped = rr_change.groupby("ownersortname_c").agg(
                {"Net Balance_p": "sum", "acctnbr": "count", "riskratingcd_c": "first"}
            ).reset_index()
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname_c"], r["riskratingcd_c"], -r["Net Balance_p"], r["acctnbr"]])

        # Paid to Zero Balance (fixed condition)
        zeroed = joined[(joined["Net Balance_c"] == 0) & (joined["Net Balance_p"] != 0)]
        if not zeroed.empty:
            rows.append(["Paid to a Zero Balance", "", "", "", ""])
            grouped = zeroed.groupby("ownersortname_c").agg(
                {"Net Balance_p": "sum", "acctnbr": "count", "riskratingcd_c": "first"}
            ).reset_index()
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname_c"], r["riskratingcd_c"], -r["Net Balance_p"], r["acctnbr"]])

        # LOC Advance
        loc = joined[joined["Net Balance_c"] > joined["Net Balance_p"]]
        if not loc.empty:
            rows.append(["LOC Advance", "", "", "", ""])
            grouped = loc.groupby("ownersortname_c").agg(
                {"Net Balance_c": "sum", "Net Balance_p": "sum", "acctnbr": "count", "riskratingcd_c": "first"}
            ).reset_index()
            grouped["delta"] = grouped["Net Balance_c"] - grouped["Net Balance_p"]
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname_c"], r["riskratingcd_c"], r["delta"], ""])

        # Get loans from current month with CO and CLS status to identify what happened
        # These loans existed in prev month but now have different status
        curr_co_cls = full_curr[full_curr["curracctstatcd"].isin(["CO", "CLS"])]
        
        # Charged Off Loans - status changed to CO
        charged_off_accts = curr_co_cls[curr_co_cls["curracctstatcd"] == "CO"]["acctnbr"].tolist()
        charged_off = p[p["acctnbr"].isin(charged_off_accts)]
        if not charged_off.empty:
            rows.append(["Charged Off Loans", "", "", "", ""])
            grouped = group_by_customer(charged_off, "ownersortname", "Net Balance")
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname"], "", -r["Net Balance"], r["acctnbr"]])
        
        # Paid Off Loans - status changed to CLS
        paid_off_accts = curr_co_cls[curr_co_cls["curracctstatcd"] == "CLS"]["acctnbr"].tolist()
        paid_off = p[p["acctnbr"].isin(paid_off_accts)]
        if not paid_off.empty:
            rows.append(["Paid Off Loans", "", "", "", ""])
            grouped = group_by_customer(paid_off, "ownersortname", "Net Balance")
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname"], "", -r["Net Balance"], r["acctnbr"]])
        
        # Other disappearances (loans that completely disappeared, not in current data at all)
        all_handled = set(charged_off_accts + paid_off_accts)
        other_disappeared = p[(~p["acctnbr"].isin(full_curr["acctnbr"])) & (~p["acctnbr"].isin(all_handled))]
        if not other_disappeared.empty:
            rows.append(["Other Closed Loans", "", "", "", ""])
            grouped = group_by_customer(other_disappeared, "ownersortname", "Net Balance")
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname"], "", -r["Net Balance"], r["acctnbr"]])

        # Partial Charge Offs
        partial = joined[joined["cobal_c"] > joined["cobal_p"]]
        if not partial.empty:
            rows.append(["Partial Charge Offs", "", "", "", ""])
            grouped = partial.groupby("ownersortname_c").agg(
                {"cobal_c": "sum", "cobal_p": "sum", "acctnbr": "count", "riskratingcd_c": "first"}
            ).reset_index()
            grouped["delta"] = -(grouped["cobal_c"] - grouped["cobal_p"])
            for _, r in grouped.iterrows():
                rows.append(["", r["ownersortname_c"], r["riskratingcd_c"], r["delta"], r["acctnbr"]])

        # --- Net Payments & Advances ---
        # Calculate sum of all other reconciliation items (excluding start/end balances and net adjustments)
        sum_other = 0
        for row in rows[2:]:  # skip header and start balance
            try:
                amt = float(row[3])
                sum_other += amt
            except Exception:
                continue

        net_adj = end_bal - start_bal
        net_payments_advances = net_adj - sum_other

        rows.append(["Net Payments & Advances", "", "", net_payments_advances, ""])
        rows.append(["Net Adjustments", "", "", net_adj, ""])
        rows.append([end_date, "", "", round(end_bal, 2), len(c)])

        return pd.DataFrame(rows)



    bold_font   = Font(bold=True)
    title_font  = Font(bold=True, size=16)

    thin_side   = Side(style="thin")
    dbl_side    = Side(style="double")

    def make_border(top=False, bottom=False, left=False, right=False, double_bottom=False) -> Border:
        """Return a Border with requested edges."""
        return Border(
            top    = thin_side if top else None,
            bottom = (dbl_side if double_bottom else thin_side) if bottom else None,
            left   = thin_side if left else None,
            right  = thin_side if right else None,
        )

    # ========================== WRITE  ==========================
    def write_out_excel():
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            ws_title, startrow = "Loan Report", 1

            # Static page header
            pd.DataFrame(
                [
                    ["BRISTOL COUNTY SAVINGS BANK"],
                    ["MONTHLY CLASSIFIED LOANS LIST"],
                    [f"DATE {curr_end}"],
                    [],
                    ["RECONCILIATION"],
                ]
            ).to_excel(writer, sheet_name= ws_title, index=False, header=False, startrow=startrow-1, startcol=1)
            startrow += 5

            recon_title_rows, recon_boxes = [], []
            summary_title_rows, header_rows, totals_rows = [], [], []

            # --------------------- 1. RECONCILIATION BOXES ---------------------
            for label, (maj, rr) in cats.items():
                # Include MLN loans with risk rating 4 in Grade 4 Commercial
                # Filter to only ACT and NPFM loans for the reconciliation categories
                if label == "4-CML":
                    prev_cat = df_prev[((df_prev["mjaccttypcd"]==maj) | 
                                       ((df_prev["mjaccttypcd"]=="MLN") & (df_prev["riskratingcd"]==rr))) & 
                                      (df_prev["riskratingcd"]==rr) & 
                                      (df_prev["curracctstatcd"].isin(["ACT", "NPFM"]))]
                    curr_cat = df_curr[((df_curr["mjaccttypcd"]==maj) | 
                                       ((df_curr["mjaccttypcd"]=="MLN") & (df_curr["riskratingcd"]==rr))) & 
                                      (df_curr["riskratingcd"]==rr) & 
                                      (df_curr["curracctstatcd"].isin(["ACT", "NPFM"]))]
                else:
                    prev_cat = df_prev[(df_prev["mjaccttypcd"]==maj)&(df_prev["riskratingcd"]==rr)&
                                      (df_prev["curracctstatcd"].isin(["ACT", "NPFM"]))]
                    curr_cat = df_curr[(df_curr["mjaccttypcd"]==maj)&(df_curr["riskratingcd"]==rr)&
                                      (df_curr["curracctstatcd"].isin(["ACT", "NPFM"]))]
                
                if prev_cat.empty and curr_cat.empty:
                    continue

                prev_date = prev_cat["effdate"].iloc[0].strftime("%m/%d/%Y") if not prev_cat.empty else "N/A"
                curr_date = curr_cat["effdate"].iloc[0].strftime("%m/%d/%Y") if not curr_cat.empty else "N/A"

                # Title row outside border
                pd.DataFrame([[f"Grade {rr} Loans ({'Commercial' if maj!='MTG' else 'Residential'})", "", "$", "#"]]) \
                    .to_excel(writer, sheet_name= ws_title, index=False, header=False, startrow=startrow, startcol=1)
                recon_title_rows.append(startrow+1)
                startrow += 1

                # Build reconciliation rows
                rows = [[prev_date, "", prev_cat["Net Balance"].sum(), len(prev_cat)]]
                reco = build_reconciliation(prev_cat, curr_cat, label)
                BAD  = {"customer name","current rr","cml grade 4","cml grade 5",
                        "mtg grade 5","substandard","special mention"}
                reco = reco[~reco[0].str.lower().fillna("").isin(BAD)]
                blank = lambda v: pd.isna(v) or str(v).strip()==""
                for _, r in reco.iterrows():
                    lbl, name, *_ , amt, cnt = r.tolist()
                    if lbl and blank(name) and blank(amt):
                        rows.append([lbl,"","",""])
                    elif str(lbl).lower().startswith("net"):
                        rows.append([lbl,"",amt,""])
                    elif name and isinstance(amt,(int,float)):
                        rows.append(["",name,amt,cnt])
                rows.append([f"{curr_date} Balance", "", curr_cat["Net Balance"].sum(), len(curr_cat)])

                pd.DataFrame(rows).to_excel(writer, sheet_name= ws_title, index=False, header=False,
                                            startrow=startrow, startcol=1)
                recon_boxes.append((startrow+1, startrow+len(rows), 2, 5))   # full border extents
                startrow += len(rows)+2

            # --------------------- 2. SUMMARY SECTIONS ---------------------
            TITLES = {"4-CML": "4 - Special Mention Commercial",
                    "5-CML": "5 - Substandard Commercial",
                    "5-MTG": "5 - Substandard Residential"}

            for label, (maj, rr) in cats.items():
                # Include MLN loans with risk rating 4 in Grade 4 Commercial
                # Only show ACT and NPFM loans in summary sections
                if label == "4-CML":
                    df = format_for_export(df_curr[((df_curr["mjaccttypcd"]==maj) | 
                                                   ((df_curr["mjaccttypcd"]=="MLN") & (df_curr["riskratingcd"]==rr))) & 
                                                  (df_curr["riskratingcd"]==rr) & 
                                                  (df_curr["curracctstatcd"].isin(["ACT", "NPFM"]))])
                else:
                    df = format_for_export(df_curr[(df_curr["mjaccttypcd"]==maj)&(df_curr["riskratingcd"]==rr)&
                                                   (df_curr["curracctstatcd"].isin(["ACT", "NPFM"]))])
                
                if df.empty: continue
                if label=="5-MTG":
                    df.rename(columns={"NAICS Description":"PROPERTY ADDRESS"}, inplace=True)

                df.drop(columns=["Risk Rating","Major"], errors="ignore", inplace=True)

                # Account Number numeric
                if "acctnbr" in df.columns:
                    df["acctnbr"] = pd.to_numeric(df["acctnbr"], errors="coerce")
                    df["acctnbr"] = df["acctnbr"].astype("Int64")

                # Sort
                dpd_col = next((c for c in df.columns if c.replace("\n"," ").strip().lower()=="days past due"), None)
                df.sort_values(["Customer Name"]+([dpd_col] if dpd_col else []),
                            ascending=[True,False] if dpd_col else [True], inplace=True)

                # Date fix
                for col in df.columns:
                    if "next" in col.lower() and "due" in col.lower():
                        df[col] = pd.to_datetime(df[col]).dt.date

                # Title row
                pd.DataFrame([[TITLES[label]]]).to_excel(writer, sheet_name= ws_title, index=False, header=False,
                                                        startrow=startrow, startcol=0)
                summary_title_rows.append(startrow+1)
                startrow+=1

                # Header row
                pd.DataFrame([df.columns]).to_excel(writer, sheet_name= ws_title, index=False, header=False,
                                                    startrow=startrow, startcol=0)
                header_rows.append((startrow+1, len(df.columns)))
                startrow+=1

                # Data
                df.to_excel(writer, sheet_name= ws_title, index=False, header=False,
                            startrow=startrow, startcol=0, na_rep="")
                startrow+=len(df)

                # Totals row (numeric sums only)
                totals=[]
                for col in df.columns:
                    norm = col.replace("\n"," ").strip().lower()
                    if norm=="customer name": totals.append("")
                    elif norm=="days past due": totals.append("")
                    elif pd.api.types.is_numeric_dtype(df[col]):
                        totals.append(df[col].sum())
                    else:
                        totals.append("")
                pd.DataFrame([totals], columns=df.columns).to_excel(writer, sheet_name= ws_title, index=False,
                                                                    header=False, startrow=startrow, startcol=0)
                totals_rows.append(startrow+1)
                startrow += 2

        # ===================================================================
        #                        OPENPYXL FORMATTING
        # ===================================================================
        wb = load_workbook(output_path)
        ws = wb["Loan Report"]

        # Static header bold
        for rr in (1,2,3,4,6): ws[f"B{rr}"].font = bold_font
        for rr in (5,6) : ws[f"B{rr}"].font = bold_font
        # Recon titles bold
        for rr in recon_title_rows:
            for cc in range(2,6): ws.cell(rr,cc).font=Font(bold=True,italic=True)

        # Recon boxes with full borders
        for top,bot,l,r in recon_boxes:
            for rr in range(top,bot+1):
                ws.cell(rr,l).font = bold_font
                if isinstance(ws.cell(rr,4).value,(int,float)):
                    ws.cell(rr,4).number_format='$#,##0.00;($#,##0.00)'
            for rr in range(top,bot+1):
                for cc in range(l,r+1):
                    ws.cell(rr,cc).border = make_border(
                        top=(rr==top or (rr==bot and cc in (4,5))),
                        bottom=(rr==bot),
                        left=(cc==l),
                        right=(cc==r)
                    )

        # Summary titles big & bold
        for rr in summary_title_rows:
            ws.cell(rr,1).font = title_font

        # Summary headers boxed & centered
        for rr,ncols in header_rows:
            hdr_map = {ws.cell(rr,c).value:c for c in range(1,ncols+1)}
            for c in range(1,ncols+1):
                h=ws.cell(rr,c)
                h.font=bold_font
                h.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
                h.border = make_border(top=True,bottom=True,left=True,right=True)

            acct = hdr_map.get("Account\nNumber")
            curr = hdr_map.get("Current\nBalance")
            chg  = hdr_map.get("Charge\nOff")
            net  = hdr_map.get("Net\nBalance")
            dpd  = hdr_map.get("Days Past\nDue")
            non  = hdr_map.get("Non\nAccrual")
            nxt  = hdr_map.get("Next\nPayment\nDue Date")

            # Find totals row
            r0=rr+1
            while ws.cell(r0,1).value not in (None,"","Total Loans:", "Total Loans"): r0+=1
            last=r0-1; tot=r0

            # Numeric formatting
            if acct:
                for row in range(rr+1,last+1):
                    c=ws.cell(row,acct)
                    if isinstance(c.value,str) and c.value.isdigit():
                        c.value=int(c.value)
                    c.number_format='0'

            for col in (curr,chg,net):
                if col:
                    for row in range(rr+1,last+1):
                        cv=ws.cell(row,col)
                        if isinstance(cv.value,(int,float)):
                            cv.number_format='#,##0.00'

            # Place Total Loans in D, sums already in E/F/G
            if curr:
                ws.cell(tot, curr-1, f"Total Loans: {last-rr}")

            # Center alignment
            center = Alignment(horizontal='center',vertical='center')
            if dpd:
                for row in range(rr+1,last+1): ws.cell(row,dpd).alignment=center
            if non:
                for row in range(rr+1,last+1): ws.cell(row,non).alignment=center
            if nxt:
                for row in range(rr+1,last+1):
                    cell=ws.cell(row,nxt)
                    if isinstance(cell.value,pd.Timestamp): cell.value=cell.value.date()
                    if isinstance(cell.value,(datetime,date)):
                        cell.number_format='mm/dd/yyyy'

        # Totals rows: bold, thin top, double bottom (no L/R)
        for t in totals_rows:
            for col in (4,5,6,7):       # D,E,F,G
                c=ws.cell(t,col)
                c.font=bold_font
                c.border = make_border(top=True,bottom=True,double_bottom=True)

            # Currency formatting for sums
            for col in (5,6,7):         # E,F,G
                if isinstance(ws.cell(t,col).value,(int,float)):
                    ws.cell(t,col).number_format='#,##0.00'

        # Column widths with caps & minimums
        CAPS = {"Next\nPayment\nDue Date":12, "Responsibility Officer":28,
                "Account\nNumber":15, "Current\nBalance":15,
                "Charge\nOff":15, "Net\nBalance":15}

        header_lookup={}
        for rr,_ in header_rows:
            for cc in range(1,ws.max_column+1):
                if (val:=ws.cell(rr,cc).value) and cc not in header_lookup:
                    header_lookup[cc]=val

        for col in ws.columns:
            idx=col[0].column
            hdr=header_lookup.get(idx,"")
            cap=CAPS.get(hdr)
            longest,numeric_only,num_len = 0,True,0
            for cell in col:
                txt=str(cell.value or "")
                longest=max(longest,*[len(x) for x in txt.split("\n")])
                if isinstance(cell.value,(int,float)):
                    num_len=max(num_len,len(txt))
                else:
                    numeric_only=False
            width=(num_len+1) if numeric_only else (longest+2)
            if hdr in {"Current\nBalance","Charge\nOff","Net\nBalance"}:
                width=max(width,15)
            if cap: width=min(width,cap) if not numeric_only else max(width,cap)
            ws.column_dimensions[get_column_letter(idx)].width = width

        wb.save(output_path)
    
    write_out_excel()
    
    print(f"Report generated: {output_path}")
    
    # Distribution
    if src.config.EMAIL_TO:
        subject = "Classifieds Report - Resolution Committee"
        body = """Hi,

Attached is the Classifieds Report for the Resolution Committee Package. If you have any questions, please reach out to BusinessIntelligence@bcsbmail.com

Thanks!"""
        
        import cdutils.distribution # type: ignore
        cdutils.distribution.email_out(
            recipients=src.config.EMAIL_TO, 
            bcc_recipients=src.config.EMAIL_CC, 
            subject=subject, 
            body=body, 
            attachment_paths=[output_path]
        )
        print(f"Email sent to {len(src.config.EMAIL_TO)} recipients")
    else:
        print("Development mode - email not sent")


if __name__ == "__main__":
    main()
    print("Complete!")